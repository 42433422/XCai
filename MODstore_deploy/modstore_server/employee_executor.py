"""AI员工执行器：基于 employee_config_v2 的真实执行管道。"""

from __future__ import annotations

import asyncio
import csv
import io
import json
import logging
import os
import time
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

import httpx

from modstore_server.runtime_async import run_coro_sync as _run_coro_sync
from modstore_server.services.llm import chat_dispatch_via_session
from modstore_server.employee_runtime import (
    build_employee_context,
    load_employee_pack,
    parse_employee_config_v2,
)
from modstore_server.models import EmployeeExecutionMetric, get_session_factory

logger = logging.getLogger(__name__)


def _get_section(config: Dict[str, Any], section: str) -> Dict[str, Any]:
    if not isinstance(config, dict):
        return {}
    if section in config and isinstance(config.get(section), dict):
        return config.get(section) or {}
    return config


def _perception_excel(input_data: Any) -> Dict[str, Any]:
    """解析 .xlsx 内容（base64 或 data URL）。"""
    import base64
    import io

    try:
        import openpyxl
    except ImportError:
        return {
            "normalized_input": input_data,
            "type": "excel",
            "parse_error": "请安装 openpyxl: pip install openpyxl",
        }

    raw = input_data
    if isinstance(input_data, dict):
        raw = input_data.get("content", input_data.get("base64", ""))
    if isinstance(raw, str) and raw.startswith("data:"):
        raw = raw.split(",", 1)[-1]
    if not raw:
        return {"normalized_input": input_data, "type": "excel", "parse_error": "empty payload"}

    try:
        wb = openpyxl.load_workbook(io.BytesIO(base64.b64decode(raw)), read_only=True, data_only=True)
        sheets_data: Dict[str, Any] = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: List[List[str]] = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(c) if c is not None else "" for c in row])
            sheets_data[sheet_name] = {
                "rows": rows,
                "row_count": len(rows),
                "col_count": len(rows[0]) if rows else 0,
            }
        wb.close()
        return {"normalized_input": sheets_data, "type": "excel", "parse_ok": True}
    except Exception as e:  # noqa: PERF203
        return {"normalized_input": input_data, "type": "excel", "parse_error": str(e)}


def _perception_image(input_data: Any, session, user_id: int) -> Dict[str, Any]:
    """优先使用多模态 LLM 描述图片（需 OpenAI 兼容 Key）。"""
    raw = input_data
    if isinstance(input_data, dict):
        raw = input_data.get("base64", input_data.get("url", input_data.get("content", "")))
    if isinstance(raw, str) and raw.startswith("data:"):
        raw = raw.split(",", 1)[-1]

    if not raw:
        return {
            "normalized_input": input_data,
            "type": "image",
            "note": "图片解析需配置 OpenAI API Key，并在 input 中提供 base64",
        }

    image_content = raw if isinstance(raw, str) and raw.startswith("data:") else f"data:image/png;base64,{raw}"

    async def _call():
        return await chat_dispatch_via_session(
            session,
            user_id,
            "openai",
            "gpt-4o-mini",
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": "请简要描述图片中的文字与关键信息。"},
                        {"type": "image_url", "image_url": {"url": image_content}},
                    ],
                }
            ],
            max_tokens=800,
        )

    try:
        result = _run_coro_sync(_call())
        if result.get("ok"):
            return {
                "normalized_input": {"description": result.get("content", ""), "type": "image"},
                "type": "image",
                "parse_ok": True,
                "method": "vision",
            }
    except Exception as e:  # noqa: PERF203
        return {"normalized_input": input_data, "type": "image", "parse_error": str(e)}
    return {"normalized_input": input_data, "type": "image", "note": "vision 调用未返回内容"}


def _memory_long_term_chroma(employee_id: str, input_data: Dict[str, Any], _cfg: Dict[str, Any]) -> Dict[str, Any]:
    """员工长期记忆：复用 ``vector_engine`` 的 PersistentClient 单例（修文件句柄泄露），
    集合保留 Chroma 默认 embedding function 以兼容历史 ``query_texts`` 写法。
    """
    query = str(input_data.get("memory_query") or input_data.get("query") or "").strip()
    if not query:
        return {"enabled": True, "memories": [], "note": "请在 input_data 中提供 memory_query 以检索长期记忆"}
    try:
        from modstore_server import vector_engine
        from modstore_server.vector_engine import VectorEngineError
    except ImportError:
        return {"enabled": True, "memories": [], "note": "请安装 chromadb: pip install chromadb"}

    try:
        client = vector_engine.get_client()
    except VectorEngineError as e:
        return {"enabled": True, "memories": [], "error": str(e)}

    coll_name = vector_engine.employee_memory_collection_name(employee_id)
    try:
        collection = client.get_or_create_collection(
            name=coll_name,
            metadata={"hnsw:space": "cosine"},
        )
    except Exception as e:  # noqa: BLE001
        return {"enabled": True, "memories": [], "error": str(e)}

    try:
        results = collection.query(query_texts=[query], n_results=5)
    except Exception as e:  # noqa: BLE001
        return {"enabled": True, "memories": [], "error": str(e)}

    documents = (results.get("documents") or [[]])[0] or []
    distances = (results.get("distances") or [[]])[0] or []
    memories = []
    for i, doc in enumerate(documents):
        dist = float(distances[i]) if i < len(distances) else 1.0
        if dist < 0.85:
            memories.append({"content": doc, "distance": dist})
    return {"enabled": True, "memories": memories, "count": len(memories)}


def _perception_real(
    config: Dict[str, Any],
    input_data: Dict[str, Any],
    session=None,
    user_id: int = 0,
) -> Dict[str, Any]:
    p_cfg = _get_section(config, "perception")
    p_type = str(p_cfg.get("type") or "text").strip().lower()
    payload = input_data or {}
    if p_type == "text":
        return {"normalized_input": payload, "type": "text"}
    if p_type == "json":
        if isinstance(payload, dict):
            return {"normalized_input": payload, "type": "json"}
        try:
            parsed = json.loads(payload) if isinstance(payload, str) else payload
            return {"normalized_input": parsed, "type": "json"}
        except Exception as e:  # noqa: PERF203
            return {"normalized_input": payload, "type": "json", "parse_error": str(e)}
    if p_type == "csv":
        raw = payload.get("content", "") if isinstance(payload, dict) else str(payload)
        try:
            reader = csv.DictReader(io.StringIO(raw))
            rows = list(reader)
            return {"normalized_input": {"rows": rows}, "type": "csv", "row_count": len(rows)}
        except Exception as e:  # noqa: PERF203
            return {"normalized_input": payload, "type": "csv", "parse_error": str(e)}
    if p_type == "excel":
        return _perception_excel(payload)
    if p_type == "image":
        return _perception_image(payload, session, user_id)
    if p_type == "document":
        return _perception_document(payload)
    if p_type in ("web_rankings", "ai_model_rankings"):
        return _perception_web_rankings(payload)
    return {"normalized_input": payload, "type": p_type}


def _perception_document(input_data: Any) -> Dict[str, Any]:
    """文档类输入：优先抽取文本字段供认知层处理。"""
    if isinstance(input_data, dict):
        text = (
            input_data.get("content")
            or input_data.get("text")
            or input_data.get("body")
            or input_data.get("markdown")
        )
        if isinstance(text, str) and text.strip():
            meta = {
                k: v
                for k, v in input_data.items()
                if k not in ("content", "text", "body", "markdown", "base64", "url")
            }
            return {
                "normalized_input": {"text": text, "meta": meta},
                "type": "document",
                "parse_ok": True,
            }
        if input_data.get("url"):
            return {
                "normalized_input": input_data,
                "type": "document",
                "note": "document.url 需宿主或后续链路拉取正文；已原样传入认知层",
            }
    return {"normalized_input": input_data, "type": "document"}


def _perception_web_rankings(input_data: Any) -> Dict[str, Any]:
    """排行榜 / 模型对比类感知：结构化包裹后由认知层推理（执行器内无实时爬虫）。"""
    payload = input_data if isinstance(input_data, dict) else {"raw": input_data}
    return {
        "normalized_input": {
            "ranking_task": True,
            "instructions": "请基于给定 payload 完成排序、对比或摘要；若信息不足请明确说明。",
            "payload": payload,
        },
        "type": "web_rankings",
    }


def _memory_real(config: Dict[str, Any], ctx: Dict[str, Any], session, user_id: int) -> Dict[str, Any]:
    mem_cfg = _get_section(config, "memory")
    employee_id = ctx["employee_id"]
    result: Dict[str, Any] = {"session": {"employee_id": employee_id}, "long_term": None}
    short_term_cfg = mem_cfg.get("short_term") or {}
    if short_term_cfg.get("enabled", True):
        q = session.query(EmployeeExecutionMetric).filter(
            EmployeeExecutionMetric.employee_id == employee_id
        )
        if user_id > 0:
            q = q.filter(EmployeeExecutionMetric.user_id == user_id)
        recent = q.order_by(EmployeeExecutionMetric.id.desc()).limit(5).all()
        result["session"]["recent_tasks"] = [
            {
                "task": r.task,
                "status": r.status,
                "created_at": r.created_at.isoformat() if r.created_at else None,
            }
            for r in recent
        ]
    long_term_cfg = mem_cfg.get("long_term") or {}
    if long_term_cfg.get("enabled", False):
        result["long_term"] = _memory_long_term_chroma(employee_id, ctx.get("input_data") or {}, long_term_cfg)
    return result


async def _cognition_real(
    config: Dict[str, Any],
    perceived: Dict[str, Any],
    memory: Dict[str, Any],
    session,
    user_id: int,
    *,
    employee_id: str = "",
    task: str = "",
    bench_llm_override: Optional[Tuple[str, str]] = None,
) -> Dict[str, Any]:
    cog_cfg = _get_section(config, "cognition")
    agent = cog_cfg.get("agent") if isinstance(cog_cfg.get("agent"), dict) else cog_cfg
    system_prompt = agent.get("system_prompt", "你是智能员工助手")
    model_cfg = agent.get("model") if isinstance(agent.get("model"), dict) else {}

    if bench_llm_override:
        provider, model_name = bench_llm_override
    else:
        provider = str(model_cfg.get("provider") or "deepseek").strip()
        model_name = str(model_cfg.get("model_name") or "deepseek-chat").strip()
    max_tokens = int(model_cfg.get("max_tokens") or 4000)
    messages = [{"role": "system", "content": system_prompt}]
    user_input = json.dumps(perceived.get("normalized_input", {}), ensure_ascii=False)
    mem_session = memory.get("session") if isinstance(memory, dict) else None
    if mem_session:
        user_input = (
            f"{user_input}\n\n[session_context]\n{json.dumps(mem_session, ensure_ascii=False)}"
        )
    messages.append({"role": "user", "content": user_input})

    knowledge_cfg = cog_cfg.get("knowledge") if isinstance(cog_cfg.get("knowledge"), dict) else {}
    rag_meta: Dict[str, Any] = {"enabled": False, "items": [], "error": ""}
    if knowledge_cfg.get("enabled"):
        try:
            from modstore_server import rag_service

            top_k = int(knowledge_cfg.get("top_k") or 6)
            min_score = float(knowledge_cfg.get("min_score") or 0.0)
            collection_ids = knowledge_cfg.get("collection_ids")
            query_text = str(task or user_input or "").strip()[:1500]
            chunks = await rag_service.retrieve(
                user_id=int(user_id or 0),
                query=query_text,
                employee_id=str(employee_id or "") or None,
                extra_collection_ids=collection_ids if isinstance(collection_ids, list) else None,
                top_k=top_k,
                min_score=min_score,
            )
            messages = rag_service.inject_rag_into_messages(messages, chunks)
            rag_meta = {
                "enabled": True,
                "items": [c.to_dict() for c in chunks],
                "count": len(chunks),
            }
        except Exception as e:  # noqa: BLE001 — RAG 失败不阻塞员工执行
            logger.warning("cognition.knowledge retrieve 失败: %s", e)
            rag_meta = {"enabled": True, "items": [], "error": str(e)}

    if bench_llm_override:
        from modstore_server.services.llm import chat_dispatch_via_platform_only
        result = await chat_dispatch_via_platform_only(provider, model_name, messages, max_tokens=max_tokens)
    else:
        result = await chat_dispatch_via_session(
            session,
            user_id,
            provider,
            model_name,
            messages,
            max_tokens=max_tokens,
        )
    if not result.get("ok"):
        err = str(result.get("error") or "llm call failed")
        if "missing api key" in err.lower():
            err = f"missing api key for provider: {provider}"
        return {
            "reasoning": "",
            "error": err,
            "input": perceived.get("normalized_input", {}),
            "memory": memory,
            "knowledge": rag_meta,
            "provider": provider,
            "model": model_name,
        }
    return {
        "reasoning": result.get("content", ""),
        "input": perceived.get("normalized_input", {}),
        "memory": memory,
        "knowledge": rag_meta,
        "provider": provider,
        "model": model_name,
        "llm_raw": result.get("raw"),
        "system_prompt": system_prompt,  # forwarded so agent runner can use it
        "_bench_platform_only": bool(bench_llm_override),
    }


def _cognition_sync(
    config: Dict[str, Any],
    perceived: Dict[str, Any],
    memory: Dict[str, Any],
    session,
    user_id: int,
    *,
    employee_id: str = "",
    task: str = "",
    bench_llm_override: Optional[Tuple[str, str]] = None,
) -> Dict[str, Any]:
    return _run_coro_sync(
        _cognition_real(
            config,
            perceived,
            memory,
            session,
            user_id,
            employee_id=employee_id,
            task=task,
            bench_llm_override=bench_llm_override,
        )
    )


def _action_wechat_notify(actions_cfg: Dict[str, Any], reasoning: Dict[str, Any], task: str) -> Dict[str, Any]:
    """企业微信机器人 Webhook。"""
    wechat_cfg = actions_cfg.get("wechat_notify") or {}
    webhook_url = str(wechat_cfg.get("webhook_url") or "").strip()
    if not webhook_url:
        return {
            "handler": "wechat_notify",
            "status": "not_configured",
            "message": "未配置 actions.wechat_notify.webhook_url",
        }
    message_type = str(wechat_cfg.get("message_type") or "text").strip()
    content = str(reasoning.get("reasoning") or "")[:2048]
    payload: Dict[str, Any] = {"msgtype": message_type}
    if message_type == "markdown":
        payload["markdown"] = {"content": f"**AI 员工通知**\n任务: {task}\n\n{content}"}
    else:
        payload["text"] = {"content": f"【AI员工】任务:{task}\n{content}"}
    try:
        resp = httpx.post(webhook_url, json=payload, timeout=10.0)
        try:
            j = resp.json()
        except Exception:
            j = {}
        if resp.status_code == 200 and int(j.get("errcode", 0)) == 0:
            return {"handler": "wechat_notify", "status": "ok"}
        return {"handler": "wechat_notify", "status": "failed", "response": resp.text[:500]}
    except Exception as e:  # noqa: PERF203
        return {"handler": "wechat_notify", "status": "error", "error": str(e)}


def _action_openapi_tool(
    actions_cfg: Dict[str, Any],
    reasoning: Dict[str, Any],
    task: str,
    employee_id: str,
    user_id: int,
) -> Dict[str, Any]:
    """通过受控 OpenAPI 连接器调用第三方 API。

    配置示例（员工 actions.openapi_tool）::

        {
          "connector_id": 12,
          "operation_id": "createIssue",
          "params": {"project": "MOD"},
          "body": {"title": "{{task}}", "body": "{{reasoning}}"},
          "headers": {"X-Trace": "ai-employee"},
          "timeout": 20
        }
    """
    cfg = actions_cfg.get("openapi_tool") or {}
    connector_id = cfg.get("connector_id")
    operation_id = cfg.get("operation_id")
    if not connector_id or not operation_id:
        return {
            "handler": "openapi_tool",
            "error": "missing connector_id or operation_id",
        }
    try:
        connector_id_int = int(connector_id)
    except (TypeError, ValueError):
        return {"handler": "openapi_tool", "error": f"invalid connector_id: {connector_id!r}"}

    def _render(value: Any) -> Any:
        if isinstance(value, str):
            return value.replace("{{reasoning}}", str(reasoning.get("reasoning") or "")).replace(
                "{{task}}", task
            ).replace("{{employee_id}}", employee_id)
        if isinstance(value, dict):
            return {str(k): _render(v) for k, v in value.items()}
        if isinstance(value, list):
            return [_render(v) for v in value]
        return value

    try:
        from modstore_server.openapi_connector_runtime import call_generated_operation
    except Exception as exc:  # noqa: BLE001
        return {"handler": "openapi_tool", "error": f"runtime unavailable: {exc}"}

    timeout = float(cfg.get("timeout") or 30)
    result = call_generated_operation(
        connector_id=connector_id_int,
        user_id=int(user_id or 0),
        operation_id=str(operation_id),
        params=_render(cfg.get("params") or {}),
        body=_render(cfg.get("body")) if cfg.get("body") is not None else None,
        headers=_render(cfg.get("headers") or {}),
        timeout=timeout,
        source="employee",
    )
    return {
        "handler": "openapi_tool",
        "connector_id": connector_id_int,
        "operation_id": operation_id,
        "ok": bool(result.get("ok")),
        "status_code": result.get("status_code"),
        "body": result.get("body"),
        "error": result.get("error") or "",
        "duration_ms": result.get("duration_ms"),
    }


def _tpl_str(s: str, reasoning: Dict[str, Any], task: str) -> str:
    rtxt = str((reasoning or {}).get("reasoning") or "")
    return (s or "").replace("{{reasoning}}", rtxt).replace("{{task}}", task or "")


def _tpl_obj(obj: Any, reasoning: Dict[str, Any], task: str) -> Any:
    if isinstance(obj, str):
        return _tpl_str(obj, reasoning, task)
    if isinstance(obj, dict):
        return {str(k): _tpl_obj(v, reasoning, task) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_tpl_obj(x, reasoning, task) for x in obj]
    return obj


def _action_fhd_business(
    actions_cfg: Dict[str, Any], reasoning: Dict[str, Any], task: str
) -> Dict[str, Any]:
    biz = actions_cfg.get("fhd_business") or {}
    base = str(
        biz.get("fhd_base_url")
        or biz.get("base_url")
        or os.environ.get("FHD_BUSINESS_BASE_URL")
        or ""
    ).strip()
    path = str(biz.get("api_path") or biz.get("path") or "").strip().lstrip("/")
    method = str(biz.get("method") or "POST").strip().upper()
    if not base:
        return {"handler": "fhd_business", "error": "missing fhd_base_url"}
    if not path:
        return {"handler": "fhd_business", "error": "missing api_path"}
    raw_body = biz.get("body")
    body: Dict[str, Any] = {}
    if isinstance(raw_body, dict):
        tb = _tpl_obj(raw_body, reasoning, task)
        body = tb if isinstance(tb, dict) else {}
    headers_in = biz.get("headers") if isinstance(biz.get("headers"), dict) else {}
    hdrs = {str(k): _tpl_str(str(v), reasoning, task) for k, v in headers_in.items()}
    key = str(biz.get("business_key") or os.environ.get("FHD_BUSINESS_API_KEY") or "").strip()
    if key:
        hdrs.setdefault("X-FHD-Business-Key", key)
    url = f"{base.rstrip('/')}/api/business/{path}"
    try:
        timeout = float(biz.get("timeout") or 30.0)
        resp = httpx.request(method, url, json=body or None, headers=hdrs, timeout=timeout)
        return {
            "handler": "fhd_business",
            "url": url,
            "status_code": resp.status_code,
            "response": (resp.text or "")[:2000],
        }
    except Exception as e:  # noqa: PERF203
        return {"handler": "fhd_business", "error": str(e), "url": url}


def _action_agent_runner(
    actions_cfg: Dict[str, Any],
    reasoning: Dict[str, Any],
    task: str,
    employee_id: str,
    user_id: int,
) -> Dict[str, Any]:
    """Dispatch the ``agent`` handler by running an EmployeeAgentRunner ReAct loop.

    Reads ``actions.agent.workspace`` to determine the project root and whether
    write tools should be available.  Falls back to the reasoning text when the
    runner is unavailable.
    """
    try:
        from modstore_server.mod_employee_agent_runner import EmployeeAgentRunner
    except ImportError as exc:
        return {
            "handler": "agent",
            "ok": False,
            "error": f"EmployeeAgentRunner 未导入: {exc}",
        }

    agent_cfg = actions_cfg.get("agent") if isinstance(actions_cfg.get("agent"), dict) else {}
    ws_cfg = agent_cfg.get("workspace") if isinstance(agent_cfg.get("workspace"), dict) else {}
    read_only = bool(ws_cfg.get("read_only", True))
    requires_root = bool(ws_cfg.get("requires_project_root", False))

    # Try to get project_root from input payload first, then from the cognition result.
    cog_input = reasoning.get("input") or {}
    project_root_raw = (
        cog_input.get("project_root")
        or cog_input.get("workspace_root")
        or (reasoning.get("input") or {}).get("project_root")
    )
    workspace_root = "."

    if project_root_raw:
        try:
            from modstore_server.integrations.vibe_adapter import ensure_within_workspace, VibePathError
            resolved = str(ensure_within_workspace(str(project_root_raw), user_id=int(user_id or 0)))
            workspace_root = resolved
        except Exception as exc:  # noqa: BLE001
            return {
                "handler": "agent",
                "ok": False,
                "error": f"project_root 路径无效: {exc}",
            }
    elif requires_root:
        return {
            "handler": "agent",
            "ok": False,
            "error": (
                "该员工需要项目根目录才能分析文件。"
                "请在 input_data 中提供 project_root 字段（例如：{'project_root': '/path/to/project'}）。"
            ),
        }

    # Build the ctx for the runner — wire up a synchronous-compatible call_llm.
    sf = get_session_factory()

    async def _agent_call_llm(messages: List[Dict[str, Any]], **kwargs) -> Dict[str, Any]:
        mt = int(kwargs.get("max_tokens") or 2048)
        temp = float(kwargs.get("temperature") or 0.2)
        # Re-use the same provider/model stored in reasoning if available.
        provider = str(reasoning.get("provider") or "deepseek")
        model = str(reasoning.get("model") or "deepseek-chat")
        if reasoning.get("_bench_platform_only"):
            from modstore_server.services.llm import chat_dispatch_via_platform_only
            return await chat_dispatch_via_platform_only(provider, model, messages, max_tokens=mt)
        with sf() as sess:
            return await chat_dispatch_via_session(sess, user_id, provider, model, messages, max_tokens=mt)

    async def _noop_http_get(url: str, **kwargs) -> Dict[str, Any]:
        return {"ok": False, "error": "agent 模式下 HTTP 工具未启用"}

    async def _noop_http_post(url: str, **kwargs) -> Dict[str, Any]:
        return {"ok": False, "error": "agent 模式下 HTTP 工具未启用"}

    ctx: Dict[str, Any] = {
        "call_llm": _agent_call_llm,
        "http_get": _noop_http_get,
        "http_post": _noop_http_post,
        "workspace_root": workspace_root,
        "employee_id": employee_id,
        "read_only": read_only,
    }

    # Extract system_prompt from reasoning config (populated by cognition layer).
    system_prompt = str(reasoning.get("system_prompt") or "").strip()
    if not system_prompt:
        # Fall back to the cognition section if available.
        cog_cfg = reasoning.get("cognition_cfg") or {}
        ag = cog_cfg.get("agent") if isinstance(cog_cfg.get("agent"), dict) else cog_cfg
        system_prompt = str(ag.get("system_prompt") or "").strip()

    runner = EmployeeAgentRunner(ctx, workspace_root=workspace_root)

    async def _run() -> Dict[str, Any]:
        return await runner.run(task, system_prompt=system_prompt)

    try:
        result = _run_coro_sync(_run())
    except Exception as exc:  # noqa: BLE001
        logger.exception("agent runner raised employee=%s", employee_id)
        return {"handler": "agent", "ok": False, "error": f"agent 执行异常: {exc}"}

    return {
        "handler": "agent",
        "ok": result.get("ok", False),
        "summary": result.get("summary") or "",
        "rounds": result.get("rounds", 0),
        "tool_calls_count": len(result.get("tool_calls") or []),
        "workspace_root": workspace_root,
        "error": result.get("error") or "",
    }


def _actions_real(
    config: Dict[str, Any],
    reasoning: Dict[str, Any],
    task: str,
    employee_id: str,
    user_id: int = 0,
) -> Dict[str, Any]:
    actions_cfg = _get_section(config, "actions")
    handlers = actions_cfg.get("handlers") or ["echo"]
    outputs: List[Dict[str, Any]] = []
    for handler in handlers:
        if handler == "echo":
            outputs.append({"handler": "echo", "output": reasoning.get("reasoning", "")})
        elif handler == "http_request":
            http_cfg = actions_cfg.get("http_request") or {}
            url = str(http_cfg.get("url") or "").strip()
            method = str(http_cfg.get("method") or "POST").strip().upper()
            headers = http_cfg.get("headers") or {}
            body_tpl = str(http_cfg.get("body") or "")
            body = body_tpl.replace("{{reasoning}}", str(reasoning.get("reasoning") or ""))
            body = body.replace("{{task}}", task)
            if not url:
                outputs.append({"handler": "http_request", "error": "missing url"})
                continue
            try:
                resp = httpx.request(method, url, headers=headers, content=body, timeout=30.0)
                outputs.append(
                    {
                        "handler": "http_request",
                        "status_code": resp.status_code,
                        "response": resp.text[:2000],
                    }
                )
            except Exception as e:  # noqa: PERF203
                outputs.append({"handler": "http_request", "error": str(e)})
        elif handler == "webhook":
            webhook_cfg = actions_cfg.get("webhook") or {}
            url = str(webhook_cfg.get("url") or "").strip()
            if not url:
                outputs.append({"handler": "webhook", "error": "missing url"})
                continue
            payload = {
                "employee_id": employee_id,
                "task": task,
                "result": reasoning.get("reasoning", ""),
            }
            try:
                resp = httpx.post(url, json=payload, timeout=30.0)
                outputs.append({"handler": "webhook", "status_code": resp.status_code})
            except Exception as e:  # noqa: PERF203
                outputs.append({"handler": "webhook", "error": str(e)})
        elif handler == "data_sync":
            target = str((actions_cfg.get("data_sync") or {}).get("target") or "log")
            if target == "log":
                logger.info(
                    "[data_sync] employee=%s task=%s result=%s",
                    employee_id,
                    task,
                    str(reasoning.get("reasoning") or "")[:500],
                )
            outputs.append({"handler": "data_sync", "target": target, "status": "ok"})
        elif handler == "wechat_notify":
            outputs.append(_action_wechat_notify(actions_cfg, reasoning, task))
        elif handler == "openapi_tool":
            outputs.append(_action_openapi_tool(actions_cfg, reasoning, task, employee_id, user_id))
        elif handler == "fhd_business":
            outputs.append(_action_fhd_business(actions_cfg, reasoning, task))
        elif handler == "voice_output":
            vo = (
                actions_cfg.get("voice_output")
                if isinstance(actions_cfg.get("voice_output"), dict)
                else {}
            )
            text = str(reasoning.get("reasoning") or "").strip()
            outputs.append(
                {
                    "handler": "voice_output",
                    "status": "pending_tts",
                    "note": "未配置 TTS 服务：返回待合成文本，可由宿主接入阿里云/讯飞/OpenAI TTS",
                    "text_preview": text[:800],
                    "provider": str(vo.get("provider") or "").strip(),
                    "voice_id": str(vo.get("voice_id") or "").strip(),
                }
            )
        elif handler == "agent":
            outputs.append(
                _action_agent_runner(actions_cfg, reasoning, task, employee_id, user_id)
            )
        elif handler == "llm_md":
            # Alias: llm_md is single-shot LLM already done via cognition; return it.
            outputs.append({"handler": "llm_md", "output": reasoning.get("reasoning", "")})
        elif handler in ("vibe_edit", "vibe_heal", "vibe_code"):
            try:
                from modstore_server.integrations.vibe_action_handlers import dispatch_vibe_handler

                vibe_out = dispatch_vibe_handler(
                    str(handler), actions_cfg, reasoning, task, employee_id, user_id
                )
            except Exception as exc:  # noqa: BLE001
                logger.exception("vibe handler dispatch failed handler=%s", handler)
                vibe_out = {"handler": str(handler), "ok": False, "error": f"dispatch error: {exc}"}
            outputs.append(vibe_out or {"handler": str(handler), "ok": False, "error": "no output"})
        else:
            outputs.append({"handler": str(handler), "error": "unknown handler"})
    return {
        "task": task,
        "handlers": handlers,
        "outputs": outputs,
        "summary": f"executed {len(outputs)} handlers",
    }


def _extract_token_count(reasoning: Dict[str, Any]) -> int:
    raw = reasoning.get("llm_raw") if isinstance(reasoning, dict) else {}
    usage = raw.get("usage") if isinstance(raw, dict) else {}
    total = usage.get("total_tokens")
    if isinstance(total, int):
        return total
    pt = usage.get("prompt_tokens")
    ct = usage.get("completion_tokens")
    return int(pt or 0) + int(ct or 0)


def execute_employee_task(
    employee_id: str,
    task: str,
    input_data: Dict[str, Any] = None,
    user_id: int = 0,
    *,
    bench_llm_override: Optional[Tuple[str, str]] = None,
) -> Dict[str, Any]:
    t0 = time.perf_counter()
    payload = input_data or {}
    sf = get_session_factory()
    with sf() as session:
        try:
            pack = load_employee_pack(session, employee_id)
            config = parse_employee_config_v2(pack.get("manifest") or {})
            ctx = build_employee_context(employee_id, payload)
            perceived = _perception_real(config.get("perception", {}), payload, session, user_id)
            memory = _memory_real(config.get("memory", {}), ctx, session, user_id)
            reasoning = _cognition_sync(
                config.get("cognition", {}),
                perceived,
                memory,
                session,
                user_id,
                employee_id=employee_id,
                task=task,
                bench_llm_override=bench_llm_override,
            )
            result = _actions_real(config.get("actions", {}), reasoning, task, employee_id, user_id)
            duration_ms = round((time.perf_counter() - t0) * 1000, 3)
            llm_tokens = _extract_token_count(reasoning)
            session.add(
                EmployeeExecutionMetric(
                    user_id=user_id,
                    employee_id=employee_id,
                    task=task,
                    status="success",
                    duration_ms=duration_ms,
                    llm_tokens=llm_tokens,
                )
            )
            session.commit()
            try:
                from modstore_server.notification_service import notify_employee_execution_done

                notify_employee_execution_done(user_id, employee_id, task, "success")
            except Exception:
                pass
            cog_err = ""
            if isinstance(reasoning, dict):
                cog_err = str(reasoning.get("error") or "").strip()
            rex = ""
            if isinstance(reasoning, dict):
                rex = str(reasoning.get("reasoning") or "").strip()[:4000]

            return {
                "employee_id": employee_id,
                "pack": {"id": pack["pack_id"], "version": pack["version"]},
                "duration_ms": duration_ms,
                "result": result,
                "executed_at": datetime.now(timezone.utc).isoformat(),
                "llm_tokens": llm_tokens,
                "cognition_error": cog_err or None,
                "reasoning_excerpt": rex or None,
            }
        except Exception as e:
            duration_ms = round((time.perf_counter() - t0) * 1000, 3)
            session.add(
                EmployeeExecutionMetric(
                    user_id=user_id,
                    employee_id=employee_id,
                    task=task,
                    status="failed",
                    duration_ms=duration_ms,
                    llm_tokens=0,
                    error=str(e),
                )
            )
            session.commit()
            try:
                from modstore_server.notification_service import notify_employee_execution_done

                if user_id:
                    notify_employee_execution_done(user_id, employee_id, task, "failed")
            except Exception:
                pass
            raise


def get_employee_status(employee_id: str) -> Dict[str, Any]:
    sf = get_session_factory()
    with sf() as session:
        rows = (
            session.query(EmployeeExecutionMetric)
            .filter(EmployeeExecutionMetric.employee_id == employee_id)
            .order_by(EmployeeExecutionMetric.id.desc())
            .limit(100)
            .all()
        )
        ok = len([r for r in rows if r.status == "success"])
        return {
            "status": "active",
            "employee_id": employee_id,
            "execution_stats": {
                "total_executions": len(rows),
                "success_count": ok,
                "failed_count": len(rows) - ok,
                "success_rate": (ok / len(rows) * 100.0) if rows else 0,
            },
            "last_execution": rows[0].created_at.isoformat() if rows else None,
        }


def list_employees() -> List[Dict[str, Any]]:
    """列出可展示的员工包：合并数据库 ``catalog_items`` 与仅存在于 ``packages.json`` 的登记。

    - ``source`` = ``catalog``：数据库中有 ``artifact=employee_pack`` 行（执行与权限以此为准）。
    - ``source`` = ``v1_catalog``：仅本地 XC catalog（``/v1/packages``）中存在，需管理员同步入库后方可稳定执行。

    同一逻辑 ``pkg_id`` 只保留一行；数据库优先于 JSON（按 :func:`~modstore_server.catalog_store.norm_pkg_id` 去重）。
    """
    from modstore_server import catalog_store
    from modstore_server.models import CatalogItem

    merged_by_norm: Dict[str, Dict[str, Any]] = {}

    sf = get_session_factory()
    with sf() as session:
        employees = session.query(CatalogItem).filter(CatalogItem.artifact == "employee_pack").all()
        for e in employees:
            nid = catalog_store.norm_pkg_id(e.pkg_id)
            if not nid:
                continue
            merged_by_norm[nid] = {
                "id": e.pkg_id,
                "name": e.name,
                "version": e.version,
                "description": e.description,
                "price": e.price,
                "industry": e.industry,
                "created_at": e.created_at.isoformat() if e.created_at else "",
                "source": "catalog",
            }

    try:
        pending_norm_raw: Dict[str, str] = {}
        for r in catalog_store.load_store().get("packages") or []:
            if not isinstance(r, dict):
                continue
            if str(r.get("artifact") or "").strip().lower() != "employee_pack":
                continue
            nid = catalog_store.norm_pkg_id(r.get("id"))
            if not nid or nid in merged_by_norm:
                continue
            rid = str(r.get("id")).strip()
            if rid:
                pending_norm_raw.setdefault(nid, rid)

        for _nid, raw_id in pending_norm_raw.items():
            versions = catalog_store.list_versions(raw_id)
            best = versions[0] if versions else None
            if not isinstance(best, dict):
                continue
            pid = str(best.get("id") or raw_id).strip()
            merged_by_norm[nid] = {
                "id": pid,
                "name": str(best.get("name") or pid),
                "version": best.get("version"),
                "description": best.get("description"),
                "price": 0.0,
                "industry": str(best.get("industry") or "通用"),
                "created_at": str(best.get("created_at") or ""),
                "source": "v1_catalog",
            }
    except Exception as ex:  # noqa: BLE001
        logger.warning("list_employees: merge packages.json failed: %s", ex)

    out = list(merged_by_norm.values())
    out.sort(key=lambda x: str(x.get("name") or x.get("id") or "").lower())
    return out
