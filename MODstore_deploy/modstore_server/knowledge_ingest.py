"""Parse uploaded knowledge files into text chunks."""

from __future__ import annotations

import csv
import io
import json
import os
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from fastapi import HTTPException


SUPPORTED_EXTENSIONS = {".txt", ".md", ".json", ".csv", ".pdf", ".docx", ".xlsx"}
MAX_UPLOAD_BYTES = int(os.environ.get("MODSTORE_KB_MAX_UPLOAD_BYTES", str(20 * 1024 * 1024)))


def _chunk_size() -> int:
    return max(400, min(int(os.environ.get("MODSTORE_KB_CHUNK_SIZE", "1000")), 4000))


def _chunk_overlap() -> int:
    return max(0, min(int(os.environ.get("MODSTORE_KB_CHUNK_OVERLAP", "120")), 800))


def _decode_text(raw: bytes) -> str:
    for enc in ("utf-8", "utf-8-sig", "gb18030"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="ignore")


def _normalize_text(text: str) -> str:
    t = (text or "").replace("\r\n", "\n").replace("\r", "\n")
    t = re.sub(r"[ \t]+", " ", t)
    t = re.sub(r"\n{3,}", "\n\n", t)
    return t.strip()


def _parse_pdf(raw: bytes) -> str:
    try:
        from pypdf import PdfReader
    except ImportError as e:
        raise HTTPException(503, "服务器未安装 pypdf，暂不能解析 PDF") from e
    reader = PdfReader(io.BytesIO(raw))
    return "\n\n".join((page.extract_text() or "").strip() for page in reader.pages)


def _parse_pdf_pages(raw: bytes) -> List[Tuple[int, str]]:
    try:
        from pypdf import PdfReader
    except ImportError as e:
        raise HTTPException(503, "服务器未安装 pypdf，暂不能解析 PDF") from e
    reader = PdfReader(io.BytesIO(raw))
    pages: List[Tuple[int, str]] = []
    for idx, page in enumerate(reader.pages, start=1):
        text = _normalize_text((page.extract_text() or "").strip())
        if text:
            pages.append((idx, text))
    return pages


def _parse_docx(raw: bytes) -> str:
    try:
        import docx
    except ImportError as e:
        raise HTTPException(503, "服务器未安装 python-docx，暂不能解析 DOCX") from e
    doc = docx.Document(io.BytesIO(raw))
    return "\n".join(p.text for p in doc.paragraphs if p.text)


def _parse_xlsx(raw: bytes) -> str:
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise HTTPException(503, "服务器未安装 openpyxl，暂不能解析 XLSX") from e
    max_rows = max(10, min(int(os.environ.get("MODSTORE_XLSX_PREVIEW_ROWS", "80")), 300))
    max_cols = max(8, min(int(os.environ.get("MODSTORE_XLSX_PREVIEW_COLS", "40")), 120))
    max_formulas = max(0, min(int(os.environ.get("MODSTORE_XLSX_MAX_FORMULAS", "80")), 500))
    max_scan_cells = max(1000, min(int(os.environ.get("MODSTORE_XLSX_SCAN_CELLS", "30000")), 200000))

    def fmt(v: Any) -> str:
        if v is None:
            return ""
        if isinstance(v, float):
            return f"{v:g}"
        return str(v).replace("\n", " ").strip()

    def cell_text(formula_value: Any, cached_value: Any) -> str:
        f = fmt(formula_value)
        c = fmt(cached_value)
        if f.startswith("="):
            return f"公式 {f}" + (f" -> {c}" if c and c != f else "")
        return c or f

    def md_cell(text: Any) -> str:
        return fmt(text).replace("|", "\\|")

    wb_values = openpyxl.load_workbook(io.BytesIO(raw), read_only=False, data_only=True)
    wb_formulas = openpyxl.load_workbook(io.BytesIO(raw), read_only=False, data_only=False)
    lines: List[str] = []
    for ws in wb_formulas.worksheets:
        vws = wb_values[ws.title]
        used_range = ws.calculate_dimension()
        lines.append(f"## Sheet: {ws.title}")
        lines.append(f"- Used range: {used_range}")
        lines.append(f"- Size: {ws.max_row or 0} rows x {ws.max_column or 0} columns")
        merged = [str(rng) for rng in getattr(ws, "merged_cells", []).ranges]
        if merged:
            lines.append(f"- Merged cells: {', '.join(merged[:40])}" + (" ..." if len(merged) > 40 else ""))

        row_limit = min(ws.max_row or 0, max_rows)
        col_limit = min(ws.max_column or 0, max_cols)
        if row_limit and col_limit:
            lines.append("")
            lines.append("### Grid preview with coordinates")
            headers = ["Row"] + [get_column_letter(c) for c in range(1, col_limit + 1)]
            lines.append("| " + " | ".join(headers) + " |")
            lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
            for r in range(1, row_limit + 1):
                cells: List[str] = []
                has_value = False
                for c in range(1, col_limit + 1):
                    fcell = ws.cell(row=r, column=c)
                    vcell = vws.cell(row=r, column=c)
                    text = cell_text(fcell.value, vcell.value)
                    if text:
                        has_value = True
                        cells.append(md_cell(f"{fcell.coordinate}={text}"))
                    else:
                        cells.append("")
                if has_value:
                    lines.append("| " + " | ".join([str(r), *cells]) + " |")
            if (ws.max_row or 0) > row_limit or (ws.max_column or 0) > col_limit:
                lines.append(f"... preview limited to {row_limit} rows x {col_limit} columns")

        if max_formulas > 0:
            formulas: List[str] = []
            scanned = 0
            for row in ws.iter_rows():
                for fcell in row:
                    scanned += 1
                    if scanned > max_scan_cells or len(formulas) >= max_formulas:
                        break
                    value = fcell.value
                    if isinstance(value, str) and value.startswith("="):
                        cached = fmt(vws[fcell.coordinate].value)
                        formulas.append(
                            f"- {fcell.coordinate}: {value}" + (f" -> cached {cached}" if cached else "")
                        )
                if scanned > max_scan_cells or len(formulas) >= max_formulas:
                    break
            if formulas:
                lines.append("")
                lines.append("### Formulas")
                lines.extend(formulas)
                if len(formulas) >= max_formulas:
                    lines.append(f"... formulas limited to {max_formulas}")
        lines.append("")
    wb_values.close()
    wb_formulas.close()
    return "\n".join(lines)


def parse_upload(filename: str, raw: bytes) -> str:
    if len(raw) > MAX_UPLOAD_BYTES:
        raise HTTPException(400, f"文件过大（>{MAX_UPLOAD_BYTES // 1024 // 1024}MB）")
    suffix = Path(filename or "").suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        raise HTTPException(400, "仅支持 .txt/.md/.json/.csv/.pdf/.docx/.xlsx")

    if suffix in {".txt", ".md"}:
        text = _decode_text(raw)
    elif suffix == ".json":
        try:
            text = json.dumps(json.loads(_decode_text(raw)), ensure_ascii=False, indent=2)
        except json.JSONDecodeError:
            text = _decode_text(raw)
    elif suffix == ".csv":
        rows = csv.reader(io.StringIO(_decode_text(raw)))
        text = "\n".join(" | ".join(cell.strip() for cell in row if cell.strip()) for row in rows)
    elif suffix == ".pdf":
        text = _parse_pdf(raw)
    elif suffix == ".docx":
        text = _parse_docx(raw)
    elif suffix == ".xlsx":
        text = _parse_xlsx(raw)
    else:
        text = ""

    normalized = _normalize_text(text)
    if len(normalized) < 10:
        raise HTTPException(400, "未能从文件中提取有效文本")
    return normalized


def chunk_text(text: str) -> List[str]:
    size = _chunk_size()
    overlap = min(_chunk_overlap(), max(0, size - 1))
    t = _normalize_text(text)
    chunks: List[str] = []
    pos = 0
    while pos < len(t):
        end = min(len(t), pos + size)
        chunk = t[pos:end].strip()
        if chunk:
            chunks.append(chunk)
        if end >= len(t):
            break
        pos = max(pos + 1, end - overlap)
    return chunks


def chunk_text_with_metadata(
    text: str,
    *,
    page_no: Optional[int] = None,
) -> Tuple[List[str], List[Dict[str, Any]]]:
    chunks = chunk_text(text)
    metas: List[Dict[str, Any]] = []
    for i, _chunk in enumerate(chunks):
        meta: Dict[str, Any] = {"chunk_in_source": i}
        if page_no is not None:
            meta["page_no"] = int(page_no)
        metas.append(meta)
    return chunks, metas


def parse_and_chunk_with_metadata(filename: str, raw: bytes) -> Tuple[str, List[str], List[Dict[str, Any]]]:
    if len(raw) > MAX_UPLOAD_BYTES:
        raise HTTPException(400, f"文件过大（>{MAX_UPLOAD_BYTES // 1024 // 1024}MB）")
    suffix = Path(filename or "").suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        raise HTTPException(400, "仅支持 .txt/.md/.json/.csv/.pdf/.docx/.xlsx")

    if suffix == ".pdf":
        pages = _parse_pdf_pages(raw)
        if not pages:
            raise HTTPException(400, "未能从 PDF 中提取有效文本")
        all_text_parts: List[str] = []
        all_chunks: List[str] = []
        all_metas: List[Dict[str, Any]] = []
        for page_no, page_text in pages:
            all_text_parts.append(f"[第 {page_no} 页]\n{page_text}")
            chunks, metas = chunk_text_with_metadata(page_text, page_no=page_no)
            all_chunks.extend(chunks)
            all_metas.extend(metas)
        return "\n\n".join(all_text_parts), all_chunks, all_metas

    text = parse_upload(filename, raw)
    chunks, metas = chunk_text_with_metadata(text)
    if not chunks:
        raise HTTPException(400, "文本分块为空")
    return text, chunks, metas


def parse_and_chunk(filename: str, raw: bytes) -> Tuple[str, List[str]]:
    text, chunks, _metas = parse_and_chunk_with_metadata(filename, raw)
    if not chunks:
        raise HTTPException(400, "文本分块为空")
    return text, chunks
