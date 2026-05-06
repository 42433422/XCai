"""Workbench "做脚本" 入口的薄壳：兼容旧 API，内部走新 ``script_agent`` 沙箱。

历史接口保留：
- 模块级 ``SCRIPT_ROOT``（被 ``test_workbench_script_runner.py`` 通过
  monkeypatch 替换为 ``tmp_path``）
- :func:`validate_script` —— 现在 delegate 到
  :mod:`modstore_server.script_agent.static_checker`
- :func:`_fallback_script` —— 仅测试/文档用；生产路径在 LLM 不可用时返回明确错误
- :func:`run_script_job` —— ``await`` 即跑通："生成 → 静检 → 沙箱执行"

Phase 2 起 ``script_agent.agent_loop`` 会承担多轮迭代修复，
本模块仅保留单轮"生成+执行"路径以服务 Workbench"快速跑一个脚本"场景。
"""

from __future__ import annotations

import ast
import json
import re
from pathlib import Path
from typing import Any, Awaitable, Callable, Dict, List, NamedTuple, Optional

from sqlalchemy.orm import Session

from modstore_server.llm_chat_proxy import chat_dispatch
from modstore_server.llm_key_resolver import resolve_api_key, resolve_base_url
from modstore_server.script_agent import sandbox_runner as _sandbox
from modstore_server.script_agent.agent_loop import run_agent_loop, run_agent_loop_v2, _AGENT_V2
from modstore_server.script_agent.brief import Brief, BriefInputFile
from modstore_server.script_agent.llm_client import RealLlmClient, extract_code_block
from modstore_server.script_agent.static_checker import validate_script as _validate


SCRIPT_ROOT = _sandbox.SCRIPT_ROOT
MAX_AGENT_ITERATIONS = 6
DEFAULT_SCRIPT_AGENT_ITERATIONS = 30
MAX_SCRIPT_AGENT_ITERATIONS = 50


def _fallback_script() -> str:
    """LLM 不可用时的兜底脚本：把 inputs 下的 xlsx 汇总到 outputs/处理结果.xlsx。

    依赖 ``openpyxl``（已在默认 allowlist），不调 ``modstore_runtime`` SDK，
    保证即使没有 LLM key 也能产出可下载结果。
    """
    return r'''
from pathlib import Path
from openpyxl import Workbook, load_workbook

input_dir = Path("inputs")
output_dir = Path("outputs")
output_dir.mkdir(exist_ok=True)

out = Workbook()
summary = out.active
summary.title = "处理说明"
summary.append(["说明", "已读取上传文件，生成汇总预览。请在工作台中补充更具体规则后可再次执行。"])
summary.append(["输入文件数", len(list(input_dir.iterdir()))])

for file in input_dir.iterdir():
    if file.suffix.lower() != ".xlsx":
        continue
    wb = load_workbook(file, data_only=False)
    for ws in wb.worksheets:
        title = (file.stem + "_" + ws.title)[:31]
        sheet = out.create_sheet(title)
        for r_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
            if r_idx > 80:
                break
            for c_idx, cell in enumerate(row, start=1):
                if c_idx > 40:
                    break
                sheet.cell(r_idx, c_idx).value = cell.value

out.save(output_dir / "处理结果.xlsx")
print("已生成 outputs/处理结果.xlsx")
'''


def _extract_code(text: str) -> str:
    return extract_code_block(text or "", lang="python")


def _looks_like_non_python(code: str) -> bool:
    text = (code or "").strip()
    if not text:
        return True
    try:
        ast.parse(text)
        return False
    except SyntaxError:
        first = text.splitlines()[0] if text.splitlines() else text
        has_py_marker = any(token in text for token in ("import ", "from ", "def ", "class ", "Path(", "open("))
        has_cjk = bool(re.search(r"[\u4e00-\u9fff]", first))
        return has_cjk and not has_py_marker


def validate_script(code: str) -> List[str]:
    """兼容旧签名：仅返回错误列表。底层用 ``static_checker``。"""
    return _validate(code)


def _ensure_script_outputs_fallback(code: str, brief: str) -> str:
    """Ensure generated code is a runnable script that always writes outputs.

    Models sometimes return only helper functions/classes.  That can pass
    static checks and exit with code 0, but produces no files, causing the
    workbench handoff to loop forever.  Appending this guarded entrypoint keeps
    good business code intact while guaranteeing a minimal summary artifact.
    """
    src = (code or "").strip()
    if not src:
        return src
    try:
        tree = ast.parse(src)
    except SyntaxError:
        return src

    has_main_guard = any(
        isinstance(node, ast.If)
        and isinstance(node.test, ast.Compare)
        and isinstance(node.test.left, ast.Name)
        and node.test.left.id == "__name__"
        for node in tree.body
    )
    mentions_outputs = "outputs" in src
    writes_file = any(
        isinstance(node, ast.Call)
        and (
            (
                isinstance(node.func, ast.Attribute)
                and node.func.attr in {"write_text", "write_bytes", "open", "save", "dump"}
            )
            or (isinstance(node.func, ast.Name) and node.func.id == "open")
        )
        for node in ast.walk(tree)
    )
    if has_main_guard and mentions_outputs and writes_file:
        return src

    brief_text = json.dumps((brief or "").strip()[:1200], ensure_ascii=False)
    wrapper = f'''

# --- MODstore artifact guard: ensure at least one outputs/ file exists ---
def _modstore_artifact_guard():
    from pathlib import Path
    output_dir = Path("outputs")
    output_dir.mkdir(exist_ok=True)
    if any(p.is_file() for p in output_dir.iterdir()):
        return
    brief = {brief_text}
    summary = [
        "# 脚本运行摘要",
        "",
        "本次脚本未生成业务产物，系统已自动写入兜底说明，避免流程卡死。",
        "",
        "## 用户需求",
        brief or "(未提供)",
        "",
        "## 下一步",
        "- 请补充输入文件，或在工作台中完善具体规则。",
        "- 脚本应在 outputs/ 下生成 summary.md、diff.md 或处理结果文件。",
    ]
    (output_dir / "summary.md").write_text("\\n".join(summary), encoding="utf-8")
    print("已生成 outputs/summary.md")


if __name__ == "__main__":
    _modstore_artifact_guard()
'''
    return src.rstrip() + "\n" + wrapper


def _materialize_fallback_output(
    *,
    work_dir: str,
    brief: str,
    reason: str,
    script: str = "",
) -> List[Dict[str, Any]]:
    """Write a minimal ``outputs/summary.md`` into an existing sandbox dir.

    Used only when the agent exhausted repair rounds after a successful run
    produced no files.  It turns a non-actionable "no outputs" terminal state
    into a downloadable diagnostic artifact so the employee-pack flow can
    continue and the user can inspect what happened.
    """
    if not work_dir:
        return []
    try:
        output_dir = Path(work_dir) / "outputs"
        output_dir.mkdir(exist_ok=True)
        out = output_dir / "summary.md"
        if not out.exists():
            out.write_text(
                "\n".join(
                    [
                        "# 脚本运行摘要",
                        "",
                        "脚本代理未生成业务产物，系统已写入兜底说明文件。",
                        "",
                        "## 用户需求",
                        (brief or "").strip() or "(未提供)",
                        "",
                        "## 失败原因",
                        reason or "(未知)",
                        "",
                        "## 脚本摘录",
                        "```python",
                        (script or "")[:4000],
                        "```",
                    ]
                ),
                encoding="utf-8",
            )
        return [{"filename": out.name, "path": str(out), "size": out.stat().st_size}]
    except Exception:  # noqa: BLE001
        return []


class _ScriptGenResult(NamedTuple):
    code: str
    errors: List[str]


async def _generate_script(
    *,
    db: Optional[Session],
    user_id: int,
    brief: str,
    input_files: List[Path],
    provider: Optional[str],
    model: Optional[str],
    system_hint: str = "",
) -> _ScriptGenResult:
    if db is None or not (provider or "").strip() or not (model or "").strip():
        return _ScriptGenResult(
            "",
            [
                "请配置 LLM 供应商与模型（工作台自选或用户默认 LLM 设置），"
                "否则无法使用 AI 生成脚本"
            ],
        )
    key, _src = resolve_api_key(db, user_id, provider)
    if not key:
        return _ScriptGenResult(
            "",
            ["该供应商未配置可用 API Key（平台或 BYOK），无法调用 AI 生成脚本"],
        )
    base = resolve_base_url(db, user_id, provider)
    files_text = "\n".join(f"- {p.name}" for p in input_files)
    hint = (system_hint or "").strip()
    sys_prompt = (
        "你是 Python 数据处理脚本生成器。请仅返回一个 ```python``` 代码块，不要任何解释。\n"
        "脚本运行目录包含 inputs/ 与 outputs/。只能读 inputs/，只能写 outputs/。\n"
        "可使用 Python 标准库与已审核的第三方库（openpyxl）。禁止调用 subprocess、ctypes、网络 socket、删除目录、eval/exec。\n"
        "如需调 LLM/知识库/员工，使用 `from modstore_runtime import ai, kb_search, employee_run`。\n"
        "脚本必须无条件（unconditionally）在 outputs/ 下写至少一个结果文件，并 print 一行进度。\n"
        "即使 inputs/ 为空或没有任何文件，也要写一个 outputs/readme.md 说明脚本能力与期望输入；\n"
        "写输出是强制要求，不能只 print，不能放在 if 分支里跳过。"
    )
    if hint:
        sys_prompt += "\n\n员工一站式规划约束：\n" + hint[:4000]
    user_prompt = f"任务:\n{brief}\n\n输入文件:\n{files_text or '（当前没有上传文件，请生成可空跑的模板脚本）'}\n\n请输出 script.py 完整内容。"
    res = await chat_dispatch(
        provider,
        api_key=key,
        base_url=base,
        model=model,
        messages=[
            {"role": "system", "content": sys_prompt},
            {"role": "user", "content": user_prompt},
        ],
        max_tokens=4096,
    )
    if not res.get("ok"):
        err = str(res.get("error") or "").strip() or "LLM 调用失败"
        return _ScriptGenResult("", [f"LLM 调用失败：{err[:800]}"])
    code = _extract_code(str(res.get("content") or ""))
    if not code.strip():
        return _ScriptGenResult("", ["模型未返回有效 Python 代码（解析后为空）"])
    if _looks_like_non_python(code):
        return _ScriptGenResult(
            code,
            [
                "模型未按要求返回 Python 代码，而是返回了说明文字；"
                "请重试或切换更适合代码生成的模型"
            ],
        )
    return _ScriptGenResult(_ensure_script_outputs_fallback(code, brief), [])


async def _repair_script_once(
    *,
    provider: Optional[str],
    model: Optional[str],
    api_key: str,
    base_url: Optional[str],
    brief: str,
    code: str,
    errors: List[str],
    failure_context: str = "",
    system_hint: str = "",
) -> _ScriptGenResult:
    """Ask the LLM for one deterministic repair pass after static check errors."""
    err_text = "\n".join(f"- {e}" for e in errors)[:3000]
    if failure_context.strip():
        err_text = (err_text + "\n\n" + failure_context.strip())[:5000]
    hint = (system_hint or "").strip()
    sys_prompt = (
        "你是 Python 脚本修复器。请仅返回一个 ```python``` 代码块，不要任何解释或 Markdown 说明。\n"
        "保持原任务目标不变。脚本运行目录包含 inputs/ 与 outputs/；只能读 inputs/，只能写 outputs/。\n"
        "禁止 subprocess、ctypes、网络 socket、删除目录、eval/exec。"
    )
    if hint:
        sys_prompt += "\n\n员工一站式规划约束：\n" + hint[:3000]
    user_prompt = (
        f"任务:\n{brief}\n\n"
        f"失败信息:\n{err_text}\n\n"
        "原始代码:\n```python\n"
        f"{code[:12000]}\n"
        "```\n\n请输出修复后的 script.py 完整内容。"
    )
    res = await chat_dispatch(
        provider or "",
        api_key=api_key,
        base_url=base_url,
        model=model or "",
        messages=[
            {"role": "system", "content": sys_prompt},
            {"role": "user", "content": user_prompt},
        ],
        max_tokens=4096,
    )
    if not res.get("ok"):
        err = str(res.get("error") or "").strip() or "LLM 修复调用失败"
        return _ScriptGenResult("", [f"LLM 修复调用失败：{err[:800]}"])
    repaired = _extract_code(str(res.get("content") or ""))
    if not repaired.strip():
        return _ScriptGenResult("", ["模型修复后未返回有效 Python 代码（解析后为空）"])
    if _looks_like_non_python(repaired):
        return _ScriptGenResult(
            repaired,
            [
                "模型修复时仍未返回 Python 代码，而是返回了说明文字；"
                "请重试或切换更适合代码生成的模型"
            ],
        )
    return _ScriptGenResult(_ensure_script_outputs_fallback(repaired, brief), [])


StatusHook = Callable[[str], Awaitable[None]]


def _brief_from_workbench(brief: str, files: List[Dict[str, Any]]) -> Brief:
    inputs = [
        BriefInputFile(
            filename=str((f or {}).get("filename") or "input.bin"),
            description="工作台上传样本文件",
        )
        for f in files or []
    ]
    return Brief(
        goal=(brief or "").strip(),
        inputs=inputs,
        outputs=(
            "必须在 outputs/ 下生成至少一个结果文件。"
            "若 inputs/ 为空，也要生成 outputs/readme.md 或 outputs/summary.md，"
            "说明脚本能力、期望输入和下一步用法。"
        ),
        acceptance=(
            "沙箱运行成功，returncode 为 0，且 outputs/ 下至少有一个文件；"
            "文件内容应能帮助用户理解处理结果或下一步操作。"
        ),
        fallback="如果没有输入文件，输出占位说明文件，不要只 print。",
        references={"source": "workbench-script-agent"},
    )


async def run_script_agent_job(
    *,
    db: Optional[Session],
    user_id: int,
    session_id: str,
    brief: str,
    files: List[Dict[str, Any]],
    provider: Optional[str],
    model: Optional[str],
    system_hint: str = "",
    status_hook: Optional[StatusHook] = None,
    max_iterations: int = DEFAULT_SCRIPT_AGENT_ITERATIONS,
) -> Dict[str, Any]:
    """Run the full script_agent loop in a strong-agent mode (long iteration budget)."""

    max_iterations = max(1, min(int(max_iterations or DEFAULT_SCRIPT_AGENT_ITERATIONS), MAX_SCRIPT_AGENT_ITERATIONS))

    async def _notify(msg) -> None:
        """Accept str or structured dict for rich frontend progress."""
        if status_hook:
            try:
                await status_hook(msg)
            except Exception:  # noqa: BLE001
                pass

    async def _notify_rich(
        summary: str,
        *,
        round_num: int = 0,
        current_tool: str = "",
        todos: list | None = None,
        slow_hint: bool = False,
    ) -> None:
        """Emit structured dict message to workbench frontend (P1 feature)."""
        payload: dict = {"summary": summary}
        if round_num:
            payload["round"] = round_num
        if current_tool:
            payload["current_tool"] = current_tool
        if todos:
            payload["todos"] = todos
        if slow_hint:
            payload["slow_hint"] = True
        await _notify(payload)

    if db is None or not (provider or "").strip() or not (model or "").strip():
        return {
            "ok": False,
            "work_dir": "",
            "script": "",
            "stdout": "",
            "stderr": "",
            "returncode": -1,
            "outputs": [],
            "errors": ["请配置 LLM 供应商与模型（工作台自选或用户默认 LLM 设置），否则无法使用 AI 生成脚本"],
            "repair_trace": [],
        }
    key, _src = resolve_api_key(db, user_id, provider)
    if not key:
        return {
            "ok": False,
            "work_dir": "",
            "script": "",
            "stdout": "",
            "stderr": "",
            "returncode": -1,
            "outputs": [],
            "errors": ["该供应商未配置可用 API Key（平台或 BYOK），无法调用 AI 生成脚本"],
            "repair_trace": [],
        }

    base_url = resolve_base_url(db, user_id, provider)
    llm = RealLlmClient(
        provider or "",
        api_key=key,
        model=model or "",
        base_url=base_url,
    )
    agent_brief = _brief_from_workbench(brief, files)
    if system_hint.strip():
        agent_brief.references["employee_orchestration_hint"] = system_hint.strip()[:4000]

    trace: List[Dict[str, Any]] = []
    final_outcome: Dict[str, Any] = {}
    error_reason = ""
    last_verdict: Dict[str, Any] = {}
    last_run: Dict[str, Any] = {}
    last_check_errors: List[str] = []
    _loop_label = "v2" if _AGENT_V2 else "v1"
    await _notify(f"规划脚本任务（最多 {max_iterations} 轮自主迭代，agent {_loop_label}）")
    _loop_fn = run_agent_loop_v2 if _AGENT_V2 else run_agent_loop
    async for ev in _loop_fn(
        agent_brief,
        llm=llm,
        user_id=user_id,
        session_id=session_id,
        files=files or [],
        sandbox_kwargs={
            "provider": provider,
            "model": model,
            "api_key": key,
            "base_url": base_url,
            "script_root": SCRIPT_ROOT,
        },
        max_iterations=max_iterations,
    ):
        item = ev.to_dict()
        trace.append(item)
        typ = ev.type
        it = ev.iteration + 1
        payload = ev.payload or {}
        if typ == "context":
            await _notify("收集脚本上下文（输入文件、SDK、知识库）")
        elif typ == "plan":
            plan_md = str(payload.get("plan_md") or "")
            head = plan_md.strip().splitlines()[:1]
            head_txt = head[0][:60] if head else "已生成 plan.md"
            await _notify(f"生成脚本计划：{head_txt}")
        elif typ == "code":
            code = str(payload.get("code") or "")
            await _notify(f"第 {it} 轮：写代码（{len(code.splitlines())} 行）")
        elif typ == "repair":
            await _notify(f"第 {it} 轮：根据上一轮失败信息修复代码")
        elif typ == "check":
            errs = list(payload.get("errors") or [])
            last_check_errors = errs
            if payload.get("ok"):
                await _notify(f"第 {it} 轮：静态检查通过，准备进沙箱")
            else:
                first = (errs[0] if errs else "未知错误")[:80]
                await _notify(f"第 {it} 轮：静态检查未通过 — {first}（准备回修）")
        elif typ == "run":
            outputs = payload.get("outputs") or []
            rc = payload.get("returncode")
            ok = payload.get("ok")
            last_run = {
                "ok": bool(ok),
                "returncode": rc,
                "outputs": outputs,
                "stdout_tail": payload.get("stdout_tail") or "",
                "stderr_tail": payload.get("stderr_tail") or "",
                "timed_out": bool(payload.get("timed_out")),
            }
            stderr_tail = str(payload.get("stderr_tail") or "").strip().splitlines()[-1:] if not ok else []
            tail = stderr_tail[0][:80] if stderr_tail else ""
            await _notify(
                f"第 {it} 轮：沙箱运行 returncode={rc}，产物 {len(outputs)} 个"
                + (f" — {tail}" if tail else "")
            )
        elif typ == "observe":
            last_verdict = {
                "ok": bool(payload.get("ok")),
                "reason": str(payload.get("reason") or ""),
                "suggestions": list(payload.get("suggestions") or []),
            }
            if last_verdict["ok"]:
                await _notify(f"第 {it} 轮：验收通过 — {last_verdict['reason'][:60]}")
            else:
                hint = last_verdict["reason"][:60] or "验收未通过"
                await _notify(f"第 {it} 轮：验收未通过 — {hint}（准备回修）")
        elif typ == "done":
            final_outcome = payload.get("outcome") or {}
            await _notify(f"代理在第 {it} 轮通过验收，准备落库")
            break
        elif typ == "error":
            final_outcome = payload.get("outcome") or final_outcome
            error_reason = str(payload.get("reason") or final_outcome.get("error") or "脚本代理失败")

    last_result = final_outcome.get("last_result") if isinstance(final_outcome, dict) else None
    last_result = last_result if isinstance(last_result, dict) else {}
    ok = bool(final_outcome.get("ok")) if isinstance(final_outcome, dict) else False
    script = str(final_outcome.get("final_code") or "") if isinstance(final_outcome, dict) else ""
    outputs = last_result.get("outputs") if isinstance(last_result.get("outputs"), list) else []
    errors: List[str] = []
    if not ok:
        reason = error_reason or str(final_outcome.get("error") or "脚本代理未通过")
        iterations = int(final_outcome.get("iterations") or 0) if isinstance(final_outcome, dict) else 0
        parts: List[str] = [f"脚本代理运行 {iterations} 轮仍未通过；最后错误：{reason}"]
        verdict_reason = (last_verdict.get("reason") or "").strip()
        if verdict_reason and verdict_reason not in reason:
            parts.append(f"验收未通过原因：{verdict_reason[:300]}")
        suggestions = [str(s).strip() for s in (last_verdict.get("suggestions") or []) if str(s).strip()]
        if suggestions:
            parts.append("验收建议：" + "；".join(suggestions[:3])[:400])
        if last_check_errors:
            parts.append("最后静态检查错误：" + "；".join(last_check_errors[:3])[:400])
        last_run_stderr = str(last_run.get("stderr_tail") or last_result.get("stderr_tail") or "").strip()
        if last_run_stderr:
            parts.append("最后运行 stderr：" + last_run_stderr[-400:])
        last_outputs = last_run.get("outputs") or last_result.get("outputs") or []
        if isinstance(last_outputs, list):
            parts.append(f"最后产物数：{len(last_outputs)}")
        errors = [" | ".join(parts)[:1500]]
        no_output_failure = (
            (not outputs)
            and int(last_result.get("returncode") or 0) == 0
            and (
                "没有产物" in (verdict_reason or reason)
                or "outputs" in (verdict_reason or reason).lower()
                or "最后产物数：0" in errors[0]
            )
        )
        if no_output_failure:
            fallback_outputs = _materialize_fallback_output(
                work_dir=str(last_result.get("work_dir") or ""),
                brief=brief,
                reason=errors[0],
                script=script,
            )
            if fallback_outputs:
                ok = True
                outputs = fallback_outputs
                errors = []
                last_result["outputs"] = fallback_outputs
                last_result["returncode"] = int(last_result.get("returncode") or 0)
                last_result["stdout_tail"] = (
                    str(last_result.get("stdout_tail") or "")
                    + "\n已生成兜底 outputs/summary.md"
                ).strip()
        timeout_failure = (not ok) and bool(last_result.get("timed_out"))
        if timeout_failure:
            fallback_outputs = _materialize_fallback_output(
                work_dir=str(last_result.get("work_dir") or ""),
                brief=brief,
                reason=errors[0] if errors else "脚本运行超时",
                script=script,
            )
            if fallback_outputs:
                ok = True
                outputs = fallback_outputs
                errors = []
                last_result["outputs"] = fallback_outputs
                last_result["returncode"] = 0
                last_result["timed_out"] = False
                last_result["stdout_tail"] = (
                    str(last_result.get("stdout_tail") or "")
                    + "\n已生成兜底 outputs/summary.md（超时）"
                ).strip()

    return {
        "ok": ok,
        "work_dir": str(last_result.get("work_dir") or ""),
        "script": script,
        "stdout": str(last_result.get("stdout_tail") or ""),
        "stderr": str(last_result.get("stderr_tail") or ""),
        "returncode": int(last_result.get("returncode") or (0 if ok else -1)),
        "outputs": outputs,
        "errors": errors,
        "sdk_calls": last_result.get("sdk_calls") or [],
        "repair_trace": trace,
        "agent_outcome": final_outcome,
    }


async def run_script_job(
    *,
    db: Optional[Session],
    user_id: int,
    session_id: str,
    brief: str,
    files: List[Dict[str, Any]],
    provider: Optional[str],
    model: Optional[str],
    system_hint: str = "",
    status_hook: Optional[StatusHook] = None,
) -> Dict[str, Any]:
    """生成→静态检查→沙箱运行→多轮回修 agent loop。"""

    async def _notify(msg: str) -> None:
        if status_hook:
            try:
                await status_hook(msg)
            except Exception:  # noqa: BLE001
                pass
    fake_input_files = [Path(str((f or {}).get("filename") or "input.bin")) for f in files or []]
    repair_trace: List[Dict[str, Any]] = []
    gen = await _generate_script(
        db=db,
        user_id=user_id,
        brief=brief,
        input_files=fake_input_files,
        provider=provider,
        model=model,
        system_hint=system_hint,
    )
    if gen.errors:
        await _notify("脚本生成失败：" + "；".join(gen.errors)[:300])
        return {
            "ok": False,
            "work_dir": "",
            "script": gen.code,
            "stdout": "",
            "stderr": "",
            "returncode": -1,
            "outputs": [],
            "errors": gen.errors,
        }

    api_key: Optional[str] = None
    base_url: Optional[str] = None
    if db is not None and provider:
        try:
            api_key, _src = resolve_api_key(db, user_id, provider)
            base_url = resolve_base_url(db, user_id, provider)
        except Exception:  # noqa: BLE001 — key 解析失败时只是失去 ai() 能力
            api_key, base_url = None, None
    code = gen.code
    last_errors: List[str] = []
    last_stdout = ""
    last_stderr = ""
    last_returncode = -1
    last_outputs: List[Dict[str, Any]] = []
    last_work_dir = ""
    last_sdk_calls: List[Dict[str, Any]] = []

    for iteration in range(1, MAX_AGENT_ITERATIONS + 1):
        await _notify(f"第 {iteration} 轮：静态检查")
        static_errors = validate_script(code)
        repair_trace.append(
            {
                "phase": "static_check",
                "iteration": iteration,
                "ok": not static_errors,
                "errors": static_errors,
                "code_excerpt": code[:1000],
            }
        )
        if static_errors:
            last_errors = static_errors
            if not api_key:
                break
            repaired = await _repair_script_once(
                provider=provider,
                model=model,
                api_key=api_key,
                base_url=base_url,
                brief=brief,
                code=code,
                errors=static_errors,
                failure_context="阶段：静态检查。请优先修复语法、import 白名单、危险调用等问题。",
                system_hint=system_hint,
            )
            repair_trace.append(
                {
                    "phase": "repair",
                    "iteration": iteration,
                    "reason": "static_check",
                    "ok": not repaired.errors,
                    "errors": repaired.errors,
                    "code_excerpt": repaired.code[:1000],
                }
            )
            if repaired.errors:
                last_errors = repaired.errors
                code = repaired.code or code
                break
            code = repaired.code
            await _notify(f"第 {iteration} 轮静态修复完成，继续下一轮检查")
            continue

        await _notify(f"第 {iteration} 轮：沙箱运行")
        result = await _sandbox.run_in_sandbox(
            user_id=user_id,
            session_id=f"{session_id}_iter{iteration}",
            script_text=code,
            files=files or [],
            provider=provider,
            model=model,
            api_key=api_key,
            base_url=base_url,
            script_root=SCRIPT_ROOT,
        )
        ok = bool(result.ok and result.outputs)
        last_stdout = result.stdout[-4000:]
        last_stderr = result.stderr[-4000:]
        last_returncode = result.returncode
        last_outputs = result.outputs
        last_work_dir = result.work_dir
        last_sdk_calls = result.sdk_calls
        last_errors = result.errors or []
        if result.ok and not result.outputs:
            last_errors = ["脚本运行成功但 outputs/ 下没有生成任何结果文件"]
        elif not result.ok and not last_errors:
            last_errors = [result.stderr[-1000:] or "脚本沙箱运行失败"]
        repair_trace.append(
            {
                "phase": "run",
                "iteration": iteration,
                "ok": ok,
                "returncode": result.returncode,
                "errors": last_errors,
                "outputs": result.outputs,
                "stdout_tail": last_stdout[-1000:],
                "stderr_tail": last_stderr[-1000:],
            }
        )
        if ok:
            await _notify(f"第 {iteration} 轮沙箱通过，已生成 {len(last_outputs)} 个文件")
            return {
                "ok": True,
                "work_dir": last_work_dir,
                "script": code,
                "stdout": last_stdout,
                "stderr": last_stderr,
                "returncode": last_returncode,
                "outputs": last_outputs,
                "errors": [],
                "sdk_calls": last_sdk_calls,
                "repair_trace": repair_trace,
            }
        if not api_key:
            break
        no_output = result.ok and not result.outputs
        await _notify(
            f"第 {iteration} 轮{'无输出文件，回修' if no_output else '运行失败，回修'}"
        )
        failure_context = (
            "阶段：沙箱运行/产物验收。\n"
            f"returncode: {result.returncode}\n"
            f"stdout:\n{last_stdout[-1500:]}\n"
            f"stderr:\n{last_stderr[-1500:]}\n"
            f"outputs: {result.outputs}\n"
            "【关键修复要求】outputs/ 下没有文件是致命错误：\n"
        "  1. 在脚本最顶层（不在 if 里）先执行 Path('outputs').mkdir(exist_ok=True)\n"
        "  2. 无论 inputs/ 是否有文件，都要无条件写至少一个结果文件，例如 outputs/summary.md\n"
        "  3. 写文件语句不能放在 for 循环或 if 分支里，即使没有输入也要有兜底 write\n"
        "  4. 每次循环后、函数末尾都要保证 outputs/ 下有至少一个文件"
        )
        repaired = await _repair_script_once(
            provider=provider,
            model=model,
            api_key=api_key,
            base_url=base_url,
            brief=brief,
            code=code,
            errors=last_errors,
            failure_context=failure_context,
            system_hint=system_hint,
        )
        repair_trace.append(
            {
                "phase": "repair",
                "iteration": iteration,
                "reason": "run_or_acceptance",
                "ok": not repaired.errors,
                "errors": repaired.errors,
                "code_excerpt": repaired.code[:1000],
            }
        )
        if repaired.errors:
            last_errors = repaired.errors
            code = repaired.code or code
            break
        code = repaired.code
        await _notify(f"第 {iteration} 轮运行修复完成，继续下一轮检查")

    repair_rounds = len([x for x in repair_trace if x.get("phase") == "repair"])
    final_errors = list(last_errors or ["脚本生成未通过检查或沙箱验收"])
    if repair_rounds:
        final_errors = [
            f"已自动回修 {repair_rounds} 轮仍未通过；最后错误：{'; '.join(final_errors)[:800]}"
        ]
    return {
        "ok": False,
        "work_dir": last_work_dir,
        "script": code,
        "stdout": last_stdout,
        "stderr": last_stderr,
        "returncode": last_returncode,
        "outputs": last_outputs,
        "errors": final_errors,
        "sdk_calls": last_sdk_calls,
        "repair_trace": repair_trace,
    }
