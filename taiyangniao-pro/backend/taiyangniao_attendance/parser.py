from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, time
from pathlib import Path
import re
from typing import Iterable

from openpyxl import load_workbook

from .header_resolver import (
    ResolvedHeader,
    resolve_daily_stats_header,
    resolve_raw_records_header,
)


DATE_RE = re.compile(r"(\d{2,4})-(\d{1,2})-(\d{1,2})")
DATETIME_FORMATS = (
    "%Y-%m-%d %H:%M",
    "%Y-%m-%d %H:%M:%S",
    "%y-%m-%d %H:%M",
    "%y-%m-%d %H:%M:%S",
    "%H:%M",
)


@dataclass
class AttendanceDayRecord:
    employee_name: str
    attendance_group: str
    department: str
    employee_no: str
    position: str
    user_id: str
    work_date: date
    shift_name: str
    source_row: int
    daily_times: list[datetime] = field(default_factory=list)
    raw_times: list[datetime] = field(default_factory=list)
    leave_hours: float = 0.0
    absent_days: float = 0.0
    late_count_hint: float = 0.0
    early_count_hint: float = 0.0
    missing_card_count: float = 0.0
    work_duration_raw: float = 0.0
    attendance_day_hint: float = 0.0
    notes: list[str] = field(default_factory=list)

    def all_punch_times(self) -> list[datetime]:
        seen: set[datetime] = set()
        merged: list[datetime] = []
        for dt in sorted(self.daily_times + self.raw_times):
            if dt not in seen:
                seen.add(dt)
                merged.append(dt)
        return merged


@dataclass
class ParsedAttendanceWorkbook:
    records: list[AttendanceDayRecord]
    rows_in: int
    month: str
    daily_header: ResolvedHeader | None = None
    raw_header: ResolvedHeader | None = None


def _cell_text(value: object) -> str:
    return str(value or "").strip()


def _safe_float(value: object) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except Exception:
        return 0.0


def dingtalk_work_hours_as_hours(raw: object) -> float:
    """钉钉「每日统计」工作时长：多为分钟整数（如 547≈9.12h）；≤24 视为已是小时。"""
    v = _safe_float(raw)
    if v <= 0:
        return 0.0
    if v <= 24.0:
        return round(v, 2)
    return round(v / 60.0, 2)


def parse_work_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _cell_text(value)
    if not text:
        return None
    match = DATE_RE.search(text)
    if not match:
        return None
    year = int(match.group(1))
    if year < 100:
        year += 2000
    month = int(match.group(2))
    day = int(match.group(3))
    try:
        return date(year, month, day)
    except ValueError:
        return None


def parse_clock_datetime(value: object, *, fallback_date: date | None = None) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, time):
        if fallback_date is None:
            return None
        return datetime.combine(fallback_date, value)
    text = _cell_text(value)
    if not text:
        return None
    for fmt in DATETIME_FORMATS:
        try:
            parsed = datetime.strptime(text, fmt)
            if fmt == "%H:%M":
                if fallback_date is None:
                    return None
                return datetime.combine(fallback_date, parsed.time())
            if parsed.year < 1900 and fallback_date is not None:
                return datetime.combine(fallback_date, parsed.time())
            return parsed
        except ValueError:
            continue
    if fallback_date is not None:
        match = re.search(r"(\d{1,2}:\d{2})", text)
        if match:
            hh, mm = match.group(1).split(":")
            return datetime.combine(fallback_date, time(int(hh), int(mm)))
    return None


def extract_month_label(value: object) -> str:
    text = _cell_text(value)
    match = re.search(r"(\d{4})-(\d{2})", text)
    if not match:
        return ""
    return f"{match.group(1)}-{match.group(2)}"


def _get(row: tuple[object, ...], idx: int | None) -> object:
    if idx is None or idx < 0 or idx >= len(row):
        return None
    return row[idx]


def _parse_raw_record_times(
    workbook_path: Path,
    *,
    hint_header_row: int = 0,
    use_llm: bool = False,
) -> tuple[dict[tuple[str, date], list[datetime]], ResolvedHeader | None]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if "原始记录" not in wb.sheetnames:
            return {}, None
        ws = wb["原始记录"]
        try:
            resolved = resolve_raw_records_header(
                ws,
                hint_header_row=hint_header_row,
                use_llm=use_llm,
            )
        except ValueError:
            return {}, None

        name_idx = resolved.col("employee_name")
        date_idx = resolved.col("work_date")
        punch_idx = resolved.col("punch_time")
        punches: dict[tuple[str, date], list[datetime]] = defaultdict(list)
        for row in ws.iter_rows(min_row=resolved.data_start_row, values_only=True):
            employee_name = _cell_text(_get(row, name_idx))
            work_date = parse_work_date(_get(row, date_idx))
            punch_dt = parse_clock_datetime(_get(row, punch_idx), fallback_date=work_date)
            if employee_name and work_date and punch_dt:
                punches[(employee_name, work_date)].append(punch_dt)
        return punches, resolved
    finally:
        wb.close()


def _iter_daily_stats_times(
    row: tuple[object, ...],
    work_date: date,
    clock_cols: list[int],
) -> Iterable[datetime]:
    for idx in clock_cols:
        if idx < len(row):
            parsed = parse_clock_datetime(row[idx], fallback_date=work_date)
            if parsed:
                yield parsed


def _sum_leave_cells(row: tuple[object, ...], leave_cols: list[int]) -> float:
    total = 0.0
    for idx in leave_cols:
        total += _safe_float(_get(row, idx))
    return total


def parse_attendance_workbook(
    input_path: str | Path,
    *,
    month: str | None = None,
    header_row: int = 0,
    use_llm: bool = False,
) -> ParsedAttendanceWorkbook:
    """解析钉钉考勤导出 xlsx。

    Parameters
    ----------
    input_path:
        钉钉导出的 xlsx 路径。
    month:
        目标月份（``YYYY-MM``），若省略则尝试从 ``每日统计`` A1 单元格抽取。
    header_row:
        前端用户填写的表头所在行（1-based）。``0`` 表示让解析器自动识别。
    use_llm:
        本地规则无法识别必需列时，是否调用项目配置的 LLM 做兜底识别。
        默认 ``False``；也可通过 ``FHD_ATTENDANCE_LLM=1`` 环境变量全局开启。
    """
    workbook_path = Path(input_path)

    raw_punches, raw_resolved = _parse_raw_record_times(
        workbook_path,
        hint_header_row=header_row,
        use_llm=use_llm,
    )

    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if "每日统计" not in wb.sheetnames:
            raise ValueError("missing 每日统计 sheet")

        ws = wb["每日统计"]
        resolved = resolve_daily_stats_header(
            ws,
            hint_header_row=header_row,
            use_llm=use_llm,
        )

        detected_month = month or extract_month_label(resolved.a1_text)

        records: list[AttendanceDayRecord] = []
        rows_in = 0

        name_idx = resolved.col("employee_name")
        date_idx = resolved.col("work_date")
        group_idx = resolved.col("attendance_group")
        dept_idx = resolved.col("department")
        no_idx = resolved.col("employee_no")
        pos_idx = resolved.col("position")
        uid_idx = resolved.col("user_id")
        shift_idx = resolved.col("shift_name")
        late_idx = resolved.col("late_count")
        early_idx = resolved.col("early_count")
        miss_on_idx = resolved.col("miss_on_duty")
        miss_off_idx = resolved.col("miss_off_duty")
        absent_idx = resolved.col("absent_days")
        wh_idx = resolved.col("work_hours")
        ad_idx = resolved.col("attendance_days")

        for row_idx, row in enumerate(
            ws.iter_rows(min_row=resolved.data_start_row, values_only=True),
            start=resolved.data_start_row,
        ):
            employee_name = _cell_text(_get(row, name_idx))
            if not employee_name:
                continue

            work_date = parse_work_date(_get(row, date_idx))
            if work_date is None:
                continue

            rows_in += 1
            record = AttendanceDayRecord(
                employee_name=employee_name,
                attendance_group=_cell_text(_get(row, group_idx)),
                department=_cell_text(_get(row, dept_idx)),
                employee_no=_cell_text(_get(row, no_idx)),
                position=_cell_text(_get(row, pos_idx)),
                user_id=_cell_text(_get(row, uid_idx)),
                work_date=work_date,
                shift_name=_cell_text(_get(row, shift_idx)),
                source_row=row_idx,
                daily_times=list(
                    _iter_daily_stats_times(row, work_date, resolved.clock_time_columns)
                ),
                raw_times=sorted(raw_punches.get((employee_name, work_date), [])),
                leave_hours=_sum_leave_cells(row, resolved.leave_columns),
                absent_days=_safe_float(_get(row, absent_idx)),
                late_count_hint=_safe_float(_get(row, late_idx)),
                early_count_hint=_safe_float(_get(row, early_idx)),
                missing_card_count=_safe_float(_get(row, miss_on_idx))
                + _safe_float(_get(row, miss_off_idx)),
                work_duration_raw=_safe_float(_get(row, wh_idx)) if wh_idx is not None else 0.0,
                attendance_day_hint=_safe_float(_get(row, ad_idx)) if ad_idx is not None else 0.0,
            )
            if record.leave_hours:
                record.notes.append(f"leave_hours={record.leave_hours:g}")
            if record.absent_days:
                record.notes.append(f"absent_days={record.absent_days:g}")
            if record.missing_card_count:
                record.notes.append(f"missing_cards={record.missing_card_count:g}")
            records.append(record)

        return ParsedAttendanceWorkbook(
            records=records,
            rows_in=rows_in,
            month=detected_month,
            daily_header=resolved,
            raw_header=raw_resolved,
        )
    finally:
        wb.close()


def distinct_people_for_detail_template(records: list[AttendanceDayRecord]) -> list[tuple[str, str, str]]:
    """(部门, 考勤组/性质, 姓名) 去重，按姓名排序；用于按数据人员重排「明细」6 行块。"""
    by_name: dict[str, tuple[str, str, str]] = {}
    for r in records:
        name = str(r.employee_name or "").strip()
        if not name:
            continue
        if name not in by_name:
            by_name[name] = (
                str(r.department or "").strip(),
                str(r.attendance_group or "").strip(),
                name,
            )
    return [by_name[k] for k in sorted(by_name.keys())]
