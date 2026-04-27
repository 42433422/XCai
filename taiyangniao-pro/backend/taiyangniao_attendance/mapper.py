from __future__ import annotations

from copy import copy
from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta
from pathlib import Path
import re
import shutil
import unicodedata
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries


DETAIL_HEADER_ROWS = 3
DETAIL_PERSON_BLOCK_ROWS = 6
# 明细模板实际列可到 85+；日×2 槽约 66 列，右侧多为合计/附注，须一并清空与规范化。
DETAIL_MAX_COL = 100
# 与 ``clear_template_blocks`` 一致：左侧姓名区不参与合计。
DETAIL_SUM_COL_START = 5
# 自该列起为模板右侧符号小计（SUMIF）等，清空/写入考勤时不得覆盖。
DETAIL_TEMPLATE_SUMMARY_BEGIN_COL = 70
# 明细右侧「序号 BP—夜班 CE:CG」等与 SUMIF 区相邻列（用于快照/粘贴排除与侧栏刷新）。
DETAIL_SIDE_SUMMARY_BP_COL = 68  # 序号
DETAIL_SIDE_SUMMARY_BQ_COL = 69  # 侧栏姓名（与块内 C 列对应行）
DETAIL_SIDE_SUMMARY_CD_COL = 82  # =BQn，供夜班公式 MATCH
DETAIL_SIDE_SUMMARY_SUMIF_START_COL = 70  # BR
DETAIL_SIDE_SUMMARY_SUMIF_END_COL = 81  # CC
DETAIL_SIDE_SUMMARY_NIGHT_COLS = (83, 84, 85)  # CE, CF, CG
# 明细表内每人块首行写入「块内数字总和」的列（须避开日×2 槽：first_sym=7 时第 31 日在 67/68）。
# 模板右侧 CH（86）起在块区内多为空。
DETAIL_ONSHEET_BLOCK_TOTAL_COL = 86
# 明细含考勤格与侧栏：模板列上常见 ``[DBNum1]``，会把数字显示成中文或 ``1.`` 等怪样；先清列样式再设单元格为 ``0.0``（如 2 → 2.0）。
DETAIL_ARABIC_NUMBER_DISPLAY = "0.0"
# ``CH``（块内数字合计）：用显式小数格式避免 ``General`` 在保存后变回 ``[DBNum1]``；与考勤格一致一位小数。
DETAIL_BLOCK_TOTAL_NUMBER_DISPLAY = "0.0"

# 考勤符号整格保留，勿把「〇」等当成数字 0 改写。
_ATT_MARK: frozenset[str] = frozenset({"√", "☆", "★", "〇", "o", "O", "¤"})

_CN_BLOCK_DIGITS: dict[str, int] = {
    "零": 0,
    "〇": 0,
    "一": 1,
    "二": 2,
    "两": 2,
    "三": 3,
    "四": 4,
    "五": 5,
    "六": 6,
    "七": 7,
    "八": 8,
    "九": 9,
}


def _chinese_block_label_to_int(text: str) -> int | None:
    """块内左侧标签常见「一…三十一」纯中文数字串 → 整数；含其它汉字则不改。"""
    if not text:
        return None
    text = unicodedata.normalize("NFKC", text).strip()
    if not text:
        return None
    allowed = set(_CN_BLOCK_DIGITS) | {"十"}
    if any(c not in allowed for c in text):
        return None
    d = _CN_BLOCK_DIGITS
    if len(text) == 1:
        if text in d:
            return d[text]
        if text == "十":
            return 10
        return None
    if len(text) == 2:
        if text[0] == "十" and text[1] in d:
            return 10 + d[text[1]]
        if text[1] == "十" and text[0] in d:
            return d[text[0]] * 10
    if len(text) == 3 and text[1] == "十" and text[0] in d and text[2] in d:
        return d[text[0]] * 10 + d[text[2]]
    return None


def _plain_cell_text(value: object) -> str:
    """与 ``str()`` 相比，显式支持 openpyxl 的富文本单元格（拼接各段，便于匹配 ``18:30记加班``）。"""
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    try:
        from openpyxl.cell.rich_text import CellRichText

        if isinstance(value, CellRichText):
            return "".join(str(part) for part in value)
    except ImportError:
        pass
    return str(value)


def _replace_line_all_cn_numerals(ls: str) -> tuple[str, bool]:
    """把一行里每一段「仅由数字用字组成」的串按最长可解析规则换成阿拉伯数字（含无换行的「二一」）。"""
    if not ls:
        return ls, False
    if ls.strip() == "〇":
        return ls, False
    if ls[0] == "〇" and len(ls.strip()) > 1:
        return ls, False
    allowed = set(_CN_BLOCK_DIGITS) | {"十"}
    out: list[str] = []
    i = 0
    changed = False
    while i < len(ls):
        if ls[i] not in allowed:
            out.append(ls[i])
            i += 1
            continue
        j = i
        while j < len(ls) and ls[j] in allowed:
            j += 1
        run = ls[i:j]
        tup: tuple[int, int] | None = None
        for take in range(len(run), 0, -1):
            prefix = run[:take]
            n = _chinese_block_label_to_int(prefix)
            if n is None:
                continue
            if n == 0 and prefix == "〇":
                break
            tup = (n, take)
            break
        if tup is None:
            out.append(ls[i])
            i += 1
            continue
        n, take = tup
        out.append(str(n))
        i += take
        changed = True
    return "".join(out), changed


def _cell_text_replace_chinese_numerals(text: str) -> str | None:
    """多行、富文本拼成的单行、或「中文数+符号」混排：替换行内所有可解析的中文数字段。"""
    text = unicodedata.normalize("NFKC", text)
    lines = text.splitlines()
    out: list[str] = []
    changed = False
    for line in lines:
        ls = line.lstrip(" \t")
        indent_len = len(line) - len(ls)
        indent = line[:indent_len]
        if not ls:
            out.append(line)
            continue
        nl, _ = _replace_line_all_cn_numerals(ls)
        new_line = indent + nl
        if new_line != line:
            changed = True
        out.append(new_line)
    if not changed:
        return None
    return "\n".join(out)


def _merge_anchor_for_cell(
    row: int,
    col: int,
    overlaps: list[tuple[int, int, int, int]],
) -> tuple[int, int]:
    """在预筛选的合并矩形列表中解析 (row,col) 所属合并的左上角。"""
    for mr1, mc1, mr2, mc2 in overlaps:
        if mr1 <= row <= mr2 and mc1 <= col <= mc2:
            return mr1, mc1
    return row, col


def _normalize_chinese_numerals_in_rect(
    ws,
    r_lo: int,
    r_hi: int,
    col_lo: int,
    col_hi: int,
) -> None:
    """矩形 [r_lo..r_hi]×[col_lo..col_hi] 内：中文数字改为阿拉伯数字（合并格写左上角）。"""
    overlaps: list[tuple[int, int, int, int]] = []
    for rng in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(rng))
        if max_row < r_lo or min_row > r_hi or max_col < col_lo or min_col > col_hi:
            continue
        overlaps.append((min_row, min_col, max_row, max_col))

    processed_merge_tops: set[tuple[int, int]] = set()
    for row in range(r_lo, r_hi + 1):
        for col in range(col_lo, col_hi + 1):
            wr, wc = _merge_anchor_for_cell(row, col, overlaps)
            key = (wr, wc)
            if key in processed_merge_tops:
                continue
            processed_merge_tops.add(key)

            v = ws.cell(wr, wc).value
            if v is None:
                continue
            if isinstance(v, (int, float)) and type(v) is not bool:
                continue
            text = _plain_cell_text(v).strip().replace("\u3000", "").strip()
            if not text:
                continue
            text = unicodedata.normalize("NFKC", text).strip()
            if not text:
                continue
            # 钉钉/模板偶发文本 ``2.``：无法当数字解析时补成 2.0
            if re.fullmatch(r"-?\d+\.", text):
                try:
                    coerced = float(text)
                except ValueError:
                    coerced = None
                if coerced is not None:
                    c = ws.cell(wr, wc)
                    c.value = coerced
                    if DETAIL_SUM_COL_START <= wc < DETAIL_TEMPLATE_SUMMARY_BEGIN_COL:
                        _force_arabic_number_format(c)
                    continue
            if text.startswith("="):
                continue
            if text in _ATT_MARK:
                continue
            n = _chinese_block_label_to_int(text)
            if n is not None:
                ws.cell(wr, wc).value = n
                continue
            mixed = _cell_text_replace_chinese_numerals(text)
            if mixed is not None:
                ws.cell(wr, wc).value = mixed


def _normalize_block_chinese_numerals(
    ws,
    block_top: int,
    col_lo: int = 1,
    col_hi: int = DETAIL_MAX_COL,
) -> None:
    """每人 6 行块内 [col_lo..col_hi]：纯中文数字改为阿拉伯数字（合并格写左上角）。

    整格为考勤符号（如请假「〇」）的不改。合并区只预筛与本块矩形相交的范围，避免全表 merged 扫描卡死。
    """
    r_hi = block_top + DETAIL_PERSON_BLOCK_ROWS - 1
    _normalize_chinese_numerals_in_rect(ws, block_top, r_hi, col_lo, col_hi)


@dataclass
class DayBandEntry:
    symbol: str
    value: float


@dataclass
class EmployeeDayTemplateData:
    work_date: date
    morning: list[DayBandEntry] = field(default_factory=list)
    afternoon: list[DayBandEntry] = field(default_factory=list)
    night: list[DayBandEntry] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)


@dataclass
class EmployeeMonthTemplateData:
    employee_name: str
    attendance_group: str
    department: str
    employee_no: str
    days: dict[int, EmployeeDayTemplateData] = field(default_factory=dict)
    normal_hours: float = 0.0
    weekday_overtime_hours: float = 0.0
    sunday_overtime_hours: float = 0.0
    leave_hours: float = 0.0
    absent_hours: float = 0.0
    late_count: float = 0.0
    early_count: float = 0.0
    warnings: list[str] = field(default_factory=list)


@dataclass
class TemplateWriteResult:
    matched_employee_count: int
    unmatched_employee_names: list[str]


@dataclass(frozen=True)
class TemplateEmployeeProfile:
    employee_name: str
    base_row: int
    department: str
    nature_text: str
    block_values: tuple[float, float, float, float]
    overtime_start: time
    # B 列「HH:MM上班」解析：首段 1h 从该时刻起，与 block_values 首格上限配合 convert._profile_blocks
    morning_work_start: time | None = None
    # B 列「大小周锚YYYY-MM-DD」：该人自己的「大周六」所在周对齐；未写则用流程规则里全局锚点
    size_week_anchor: date | None = None


_NATURE_OT_RE = re.compile(
    r"(?P<h>\d{1,2})\s*[:：]\s*(?P<m>\d{2})\s*(?:记加班|加班)"
)
_NATURE_WORK_RE = re.compile(r"(?P<h>\d{1,2})\s*[:：]\s*(?P<m>\d{2})\s*上班")
_SIZE_WEEK_ANCHOR_RE = re.compile(
    r"(?:大小周锚|锚周六|个人大周六)\s*[:：]?\s*(\d{4})-(\d{1,2})-(\d{1,2})"
)


def _floor_to_saturday(d: date) -> date:
    off = (d.weekday() - 5) % 7
    return d - timedelta(days=off)


def _parse_size_week_anchor(nature_plain: str) -> date | None:
    m = _SIZE_WEEK_ANCHOR_RE.search(nature_plain)
    if not m:
        return None
    try:
        y, mo, da = int(m.group(1)), int(m.group(2)), int(m.group(3))
        return _floor_to_saturday(date(y, mo, da))
    except ValueError:
        return None


def _safe_hhmm(h: int, m: int) -> time | None:
    if 0 <= h <= 23 and 0 <= m <= 59:
        return time(h, m)
    return None


def _parse_profile_rules_from_nature_plain(nature_plain: str) -> tuple[time, tuple[float, float, float, float], time | None]:
    """从明细 B 列（规格/备注）纯文本解析：加班起算点、上午块、自定义上班点。"""
    overtime_start = time(18, 0)
    block_values: tuple[float, float, float, float] = (2.0, 2.0, 2.0, 2.0)
    morning_work_start: time | None = None

    for m in _NATURE_OT_RE.finditer(nature_plain):
        t = _safe_hhmm(int(m.group("h")), int(m.group("m")))
        if t is not None:
            overtime_start = t

    wm = _NATURE_WORK_RE.search(nature_plain)
    if wm:
        t = _safe_hhmm(int(wm.group("h")), int(wm.group("m")))
        if t is not None:
            morning_work_start = t
            block_values = (1.0, 2.0, 2.0, 2.0)
    elif "09:00" in nature_plain:
        # 兼容旧模板仅写 ``09:00``、未写「上班」二字
        morning_work_start = time(9, 0)
        block_values = (1.0, 2.0, 2.0, 2.0)

    return overtime_start, block_values, morning_work_start


def _round_display(value: float) -> float:
    """一律保留一位小数的 float（如 2.0），避免整数在列格式下被显示成 ``2.``。"""
    return round(float(value), 1)


def _ensure_template_workbook(output_path: Path, template_path: Path | None = None):
    if template_path is not None:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        if template_path.resolve() != output_path.resolve():
            shutil.copy2(template_path, output_path)
            return load_workbook(output_path)
    if output_path.exists():
        return load_workbook(output_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "明细"
    ws["A1"] = "未提供模板，已生成基础明细页"
    return wb


def open_output_workbook(output_path: Path, template_path: Path | None = None):
    return _ensure_template_workbook(output_path, template_path)


def _snapshot_first_person_block(ws, block_top: int) -> tuple[dict[tuple[int, int], object], list[tuple[int, int, int, int]]]:
    """复制首个人员块（6 行）的单元格值与合并区域（相对块内行 1-6）。

    不含 ``DETAIL_TEMPLATE_SUMMARY_BEGIN_COL`` 及以右列，避免每人块粘贴时覆盖侧栏 SUMIF/夜班公式。
    """
    cap = DETAIL_TEMPLATE_SUMMARY_BEGIN_COL - 1
    vals: dict[tuple[int, int], object] = {}
    for dr in range(DETAIL_PERSON_BLOCK_ROWS):
        for c in range(1, cap + 1):
            vals[(dr + 1, c)] = ws.cell(block_top + dr, c).value
    rel_merges: list[tuple[int, int, int, int]] = []
    for rng in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = range_boundaries(str(rng))
        if (
            min_row >= block_top
            and max_row <= block_top + DETAIL_PERSON_BLOCK_ROWS - 1
            and min_col >= 1
            and max_col <= cap
        ):
            rel_merges.append(
                (min_col, min_row - block_top + 1, max_col, max_row - block_top + 1)
            )
    return vals, rel_merges


# (dr, c) 与 ``proto_vals`` 相同：dr 为块内 1..6 行，c 为列号。
BlockCellStyle = tuple[object | None, object | None, object | None]


def _snapshot_block_cell_styles(ws, block_top: int) -> dict[tuple[int, int], BlockCellStyle]:
    """快照首块 6×DETAIL_MAX_COL 的字体、边框、对齐（供粘贴、清空与写入后恢复版式）。"""
    styles: dict[tuple[int, int], BlockCellStyle] = {}
    for dr in range(DETAIL_PERSON_BLOCK_ROWS):
        for c in range(1, DETAIL_MAX_COL + 1):
            cell = ws.cell(block_top + dr, c)
            try:
                font = copy(cell.font)
                border = copy(cell.border)
                alignment = copy(cell.alignment)
            except (TypeError, ValueError):
                font = border = alignment = None
            styles[(dr + 1, c)] = (font, border, alignment)
    return styles


def _apply_style_bundle(cell, bundle: BlockCellStyle | None) -> None:
    if not bundle:
        return
    font, border, alignment = bundle
    if font is not None:
        cell.font = font
    if border is not None:
        cell.border = border
    if alignment is not None:
        cell.alignment = alignment


def _force_arabic_number_format(cell, value: object | None = None) -> None:
    """单元格数字显示为常规阿拉伯数字（须配合列级 ``_strip_dbnum_column_styles``）。"""
    cell.number_format = DETAIL_ARABIC_NUMBER_DISPLAY


def _strip_dbnum_column_styles(ws, col_lo: int, col_hi: int) -> None:
    """去掉 ``column_dimensions`` 上的 ``[DBNum1]``，否则保存后仍按列样式显示中文数或 ``2.`` 等。"""
    for c in range(col_lo, col_hi + 1):
        letter = get_column_letter(c)
        dim = ws.column_dimensions.get(letter)
        if dim is None:
            continue
        fmt = str(dim.number_format or "")
        if "DBNum" in fmt or "dbnum" in fmt.lower():
            dim.number_format = DETAIL_ARABIC_NUMBER_DISPLAY


def _paste_one_person_block(
    ws,
    block_top: int,
    proto_vals: dict[tuple[int, int], object],
    rel_merges: list[tuple[int, int, int, int]],
    department: str,
    nature: str,
    name: str,
    proto_styles: dict[tuple[int, int], BlockCellStyle] | None = None,
) -> None:
    for (dr, c), v in proto_vals.items():
        if c >= DETAIL_TEMPLATE_SUMMARY_BEGIN_COL:
            continue
        tgt = ws.cell(block_top + dr - 1, c)
        tgt.value = v
        if proto_styles:
            st = proto_styles.get((dr, c))
            if st:
                _apply_style_bundle(tgt, st)
    for dr in range(DETAIL_PERSON_BLOCK_ROWS):
        for c in range(5, DETAIL_TEMPLATE_SUMMARY_BEGIN_COL):
            tgt = ws.cell(block_top + dr, c)
            tgt.value = None
            if proto_styles:
                st = proto_styles.get((dr + 1, c))
                if st:
                    _apply_style_bundle(tgt, st)
    ws.cell(block_top, 1).value = department
    ws.cell(block_top, 2).value = nature
    ws.cell(block_top, 3).value = name
    if proto_styles:
        for lbl_r, lbl_c in ((1, 1), (1, 2), (1, 3)):
            st = proto_styles.get((lbl_r, lbl_c))
            if st:
                _apply_style_bundle(ws.cell(block_top + lbl_r - 1, lbl_c), st)
    for min_c, r1, max_c, r2 in rel_merges:
        ref = (
            f"{get_column_letter(min_c)}{block_top + r1 - 1}:"
            f"{get_column_letter(max_c)}{block_top + r2 - 1}"
        )
        try:
            ws.merge_cells(ref)
        except ValueError:
            pass
    _normalize_block_chinese_numerals(
        ws, block_top, 1, DETAIL_TEMPLATE_SUMMARY_BEGIN_COL - 1
    )


def rebuild_detail_sheet_person_blocks(
    ws,
    people: list[tuple[str, str, str]],
    *,
    header_rows: int = DETAIL_HEADER_ROWS,
    prototype_block_top: int = 4,
) -> None:
    """按数据人员名单重排「明细」：保留前 ``header_rows`` 行与首块版式（含合并），每人仍占 6 行；打卡区清空待写。"""
    if not people:
        return
    if ws.max_row < prototype_block_top + DETAIL_PERSON_BLOCK_ROWS - 1:
        raise ValueError("明细表过短：缺少模板首个人员 6 行块")

    proto_vals, rel_merges = _snapshot_first_person_block(ws, prototype_block_top)
    proto_styles = _snapshot_block_cell_styles(ws, prototype_block_top)
    last = ws.max_row
    if last > header_rows:
        ws.delete_rows(header_rows + 1, last - header_rows)

    for i, (dept, nature, name) in enumerate(people):
        top = header_rows + 1 + i * DETAIL_PERSON_BLOCK_ROWS
        _paste_one_person_block(ws, top, proto_vals, rel_merges, dept, nature, name, proto_styles)


def find_template_base_rows(ws) -> dict[str, int]:
    """每人块仅认第 1 行（块首）C 列姓名，与 ``rebuild_detail_sheet_person_blocks`` 的 6 行步进一致。"""
    mapping: dict[str, int] = {}
    row = DETAIL_HEADER_ROWS + 1
    while row <= ws.max_row:
        name = ws.cell(row, 3).value
        if name not in (None, ""):
            key = str(name).strip()
            if key:
                mapping[key] = row
        row += DETAIL_PERSON_BLOCK_ROWS
    return mapping


def build_template_profiles(ws) -> dict[str, TemplateEmployeeProfile]:
    profiles: dict[str, TemplateEmployeeProfile] = {}
    first_block = DETAIL_HEADER_ROWS + 1
    for base_row in range(first_block, ws.max_row + 1, DETAIL_PERSON_BLOCK_ROWS):
        name = ws.cell(base_row, 3).value
        if name in (None, ""):
            continue
        employee_name = str(name).strip()
        nature_raw = ws.cell(base_row, 2).value
        nature_plain = unicodedata.normalize("NFKC", _plain_cell_text(nature_raw)).strip()
        overtime_start, block_values, morning_work_start = _parse_profile_rules_from_nature_plain(
            nature_plain
        )
        size_week_anchor = _parse_size_week_anchor(nature_plain)
        profiles[employee_name] = TemplateEmployeeProfile(
            employee_name=employee_name,
            base_row=base_row,
            department=str(ws.cell(base_row, 1).value or "").strip(),
            nature_text=nature_plain,
            block_values=block_values,
            overtime_start=overtime_start,
            morning_work_start=morning_work_start,
            size_week_anchor=size_week_anchor,
        )
    return profiles


def set_template_month(ws, month_label: str) -> None:
    if not month_label:
        return
    year_str, month_str = month_label.split("-", 1)
    ws["M1"] = int(year_str)
    ws["S1"] = int(month_str)


def _detail_calendar_anchor_col(ws, header_rows: int = DETAIL_HEADER_ROWS) -> int | None:
    """表头行中「1 日」所在列（上午格），作为日×2 槽起点；与 ``DETAIL_TEMPLATE_SUMMARY_BEGIN_COL`` 之前扫描。"""
    row = header_rows
    hi = min(DETAIL_TEMPLATE_SUMMARY_BEGIN_COL - 1, DETAIL_MAX_COL)
    for c in range(5, hi + 1):
        v = ws.cell(row, c).value
        if v in (1, "1", "１"):
            return c
        if isinstance(v, str) and unicodedata.normalize("NFKC", v).strip() in {"1", "１"}:
            return c
    return None


def _first_attendance_symbol_col(ws, sample_row: int) -> int:
    """日×2 考勤写入起始列：优先用表头「1 日」列，避免首格为空时误把第 2 日当作起点（会侵占 BP 侧栏）。

    否则在块首行从左向右找第一个考勤符号列；再回退 5。
    """
    anchor = _detail_calendar_anchor_col(ws, DETAIL_HEADER_ROWS)
    if anchor is not None:
        return int(anchor)
    for c in range(5, DETAIL_MAX_COL + 1):
        v = ws.cell(sample_row, c).value
        text = _plain_cell_text(v).strip()
        if not text:
            continue
        text = unicodedata.normalize("NFKC", text).strip()
        if text in _ATT_MARK:
            return c
    return 5


def clear_template_blocks(
    ws,
    base_rows: Iterable[int],
    proto_styles: dict[tuple[int, int], BlockCellStyle] | None = None,
) -> None:
    for base_row in base_rows:
        for row_idx in range(base_row, base_row + 6):
            for col_idx in range(5, DETAIL_TEMPLATE_SUMMARY_BEGIN_COL):
                cell = ws.cell(row_idx, col_idx)
                cell.value = None
                if proto_styles:
                    rel_r = row_idx - base_row + 1
                    st = proto_styles.get((rel_r, col_idx))
                    if st:
                        _apply_style_bundle(cell, st)


# 月度统计可链到明细侧栏 SUMIF 的列：顺序为「更具体的文案优先」，避免「平常加班」含「正常」子串误匹配。
_DETAIL_MONTH_LINK_RULES: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("星期天加班", ("星期天加班", "周日加班", "星期天", "周日", "★")),
    ("平常加班", ("平常加班", "平时加班", "平日加班", "☆")),
    ("正常上班", ("正常上班", "正班工时", "正班", "√")),
    ("请假", ("请假", "事假", "病假", "调休", "年假", "○", "〇")),
    ("旷工", ("旷工", "¤")),
    ("迟到", ("迟到",)),
    ("早退", ("早退",)),
    ("警告", ("警告",)),
)


def _detail_side_month_link_column_map(ws) -> dict[str, int]:
    """从明细表 BR..CC 表头解析「月度统计」各指标应对应的侧栏 SUMIF 列号。

    太阳鸟模板常在第 2 行放图示（与 ``_refresh_detail_side_summary_formulas`` 的 ``crit_row`` 一致），
    亦尝试第 3 行与第 1 行，避免模板改版后解析失败导致月度统计退化为静态数。
    """
    crit_rows = (1, max(1, DETAIL_HEADER_ROWS - 1), DETAIL_HEADER_ROWS)
    out: dict[str, int] = {}
    for crit_row in crit_rows:
        for c in range(DETAIL_SIDE_SUMMARY_SUMIF_START_COL, DETAIL_SIDE_SUMMARY_SUMIF_END_COL + 1):
            raw = ws.cell(crit_row, c).value
            t = unicodedata.normalize("NFKC", _plain_cell_text(raw)).replace(" ", "").replace("\n", "")
            if not t:
                continue
            for key, needles in _DETAIL_MONTH_LINK_RULES:
                if key in out:
                    continue
                if any(n in t for n in needles):
                    out[key] = c
                    break
    return out


def _detail_side_metric_symbol_columns(ws) -> tuple[int | None, int | None, int | None]:
    """兼容旧调用：从 ``_detail_side_month_link_column_map`` 取前三项。"""
    m = _detail_side_month_link_column_map(ws)
    return m.get("正常上班"), m.get("平常加班"), m.get("星期天加班")


def _refresh_detail_side_summary_formulas(ws, *, header_rows: int = DETAIL_HEADER_ROWS) -> None:
    """重写明细右侧 BP—CG：按表头符号行（图示）做 SUMIF，对每人 6 行考勤区求相邻数值之和。

    条件列引用 ``BR..CC`` 第 ``header_rows-1`` 行（太阳鸟模板为第 2 行图示：√☆★…☆〓★〓），
    数据区仍为 ``OFFSET($E$4:$BM$9, off,)`` 与 ``OFFSET($F$4:$BN$9, off,)``，
    ``off = 该人块首行行号 - (表头行数+1)``（每人仅块首一行写侧栏公式，其余行清空）。

    另解决：rebuild 复制首块侧栏 ROWS 错位、清空擦掉 BQ、日历起点列误写 BP 等问题。
    """
    first_body = header_rows + 1
    crit_row = max(1, header_rows - 1)
    mapping = find_template_base_rows(ws)
    if not mapping:
        return

    ce0 = ws.cell(first_body, DETAIL_SIDE_SUMMARY_NIGHT_COLS[0]).value
    cf0 = ws.cell(first_body, DETAIL_SIDE_SUMMARY_NIGHT_COLS[1]).value
    cg0 = ws.cell(first_body, DETAIL_SIDE_SUMMARY_NIGHT_COLS[2]).value

    side_lo = DETAIL_SIDE_SUMMARY_BP_COL
    side_hi = max(DETAIL_SIDE_SUMMARY_NIGHT_COLS)

    for base_row in sorted(set(mapping.values())):
        for rr in range(base_row + 1, base_row + DETAIL_PERSON_BLOCK_ROWS):
            for c in range(side_lo, side_hi + 1):
                ws.cell(rr, c).value = None

    for pidx, base_row in enumerate(sorted(mapping.values()), start=1):
        r = base_row
        off = r - first_body
        c_bp = ws.cell(r, DETAIL_SIDE_SUMMARY_BP_COL)
        c_bp.value = float(pidx)
        _force_arabic_number_format(c_bp, float(pidx))

        nm = ws.cell(r, 3).value
        ws.cell(r, DETAIL_SIDE_SUMMARY_BQ_COL).value = nm if nm not in (None, "") else None

        c_cd = ws.cell(r, DETAIL_SIDE_SUMMARY_CD_COL)
        c_cd.value = f"=BQ{r}"
        _force_arabic_number_format(c_cd, None)

        for c in range(DETAIL_SIDE_SUMMARY_SUMIF_START_COL, DETAIL_SIDE_SUMMARY_SUMIF_END_COL + 1):
            letter = get_column_letter(c)
            c_sum = ws.cell(r, c)
            c_sum.value = (
                f"=SUMIF(OFFSET($E$4:$BM$9,{off},),{letter}${crit_row},"
                f"OFFSET($F$4:$BN$9,{off},))"
            )
            _force_arabic_number_format(c_sum, None)

        if isinstance(ce0, str) and ce0.startswith("="):
            n = 1 + (r - first_body)

            def _night(tpl: str) -> str:
                out = re.sub(r"\$CD\d+", f"$CD{r}", tpl)
                return re.sub(r"COLUMN\(([B-D])\d+\)", lambda m: f"COLUMN({m.group(1)}{n})", out)

            c_ce = ws.cell(r, DETAIL_SIDE_SUMMARY_NIGHT_COLS[0])
            c_ce.value = _night(ce0)
            _force_arabic_number_format(c_ce, None)
            if isinstance(cf0, str) and cf0.startswith("="):
                c_cf = ws.cell(r, DETAIL_SIDE_SUMMARY_NIGHT_COLS[1])
                c_cf.value = _night(cf0)
                _force_arabic_number_format(c_cf, None)
            if isinstance(cg0, str) and cg0.startswith("="):
                c_cg = ws.cell(r, DETAIL_SIDE_SUMMARY_NIGHT_COLS[2])
                c_cg.value = _night(cg0)
                _force_arabic_number_format(c_cg, None)


def _detail_numeric_addend(value: object) -> float | None:
    """仅统计 int/float（排除 bool、日期时间、公式串等）。"""
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, (date, datetime, time)):
        return None
    try:
        from decimal import Decimal

        if isinstance(value, Decimal):
            return float(value)
    except ImportError:
        pass
    return None


def _sum_person_block_numeric_cells(
    ws,
    top_row: int,
    *,
    start_col: int = DETAIL_SUM_COL_START,
    end_col: int = DETAIL_TEMPLATE_SUMMARY_BEGIN_COL - 1,
) -> float:
    total = 0.0
    for dr in range(DETAIL_PERSON_BLOCK_ROWS):
        r = top_row + dr
        for c in range(start_col, end_col + 1):
            add = _detail_numeric_addend(ws.cell(r, c).value)
            if add is not None:
                total += add
    return total


def _ensure_onsheet_block_total_header(ws) -> None:
    """在明细第 2 行标注块内数字总和列（若该格为空）。"""
    cell = ws.cell(2, DETAIL_ONSHEET_BLOCK_TOTAL_COL)
    if cell.value in (None, ""):
        cell.value = "块内数字计"


def _write_entries(
    ws,
    row_idx: int,
    symbol_col: int,
    entries: list[DayBandEntry],
    *,
    block_top: int | None = None,
    proto_styles: dict[tuple[int, int], BlockCellStyle] | None = None,
) -> None:
    for offset, entry in enumerate(entries[:2]):
        target_row = row_idx + offset
        for col, val in (
            (symbol_col, entry.symbol),
            (symbol_col + 1, _round_display(entry.value)),
        ):
            cell = ws.cell(target_row, col)
            cell.value = val
            if proto_styles is not None and block_top is not None:
                rel_r = target_row - block_top + 1
                st = proto_styles.get((rel_r, col))
                if st:
                    _apply_style_bundle(cell, st)
            if col == symbol_col + 1 and isinstance(val, (int, float)) and type(val) is not bool:
                _force_arabic_number_format(cell)


def write_detail_sheet(
    workbook,
    employees: dict[str, EmployeeMonthTemplateData],
    *,
    month_label: str,
) -> TemplateWriteResult:
    ws = workbook["明细"] if "明细" in workbook.sheetnames else workbook.active
    set_template_month(ws, month_label)

    base_rows = find_template_base_rows(ws)
    probe_row = min(base_rows.values()) if base_rows else (DETAIL_HEADER_ROWS + 1)
    first_sym_col = _first_attendance_symbol_col(ws, probe_row)
    proto_styles_body = _snapshot_block_cell_styles(ws, probe_row)

    clear_template_blocks(ws, base_rows.values(), proto_styles_body)
    # 考勤区 E/AJ/BO… 与侧栏 BP—CH 等列常带 [DBNum1]，会盖过单元格格式导致 ``1.`` / 中文数字。
    _strip_dbnum_column_styles(ws, DETAIL_SUM_COL_START, DETAIL_ONSHEET_BLOCK_TOTAL_COL)
    for br in base_rows.values():
        _normalize_block_chinese_numerals(
            ws, br, 1, DETAIL_TEMPLATE_SUMMARY_BEGIN_COL - 1
        )

    matched = 0
    unmatched: list[str] = []

    for employee_name, payload in employees.items():
        base_row = base_rows.get(employee_name)
        if base_row is None:
            unmatched.append(employee_name)
            continue
        matched += 1
        for day, day_payload in payload.days.items():
            symbol_col = first_sym_col + (day - 1) * 2
            _write_entries(
                ws,
                base_row,
                symbol_col,
                day_payload.morning,
                block_top=base_row,
                proto_styles=proto_styles_body,
            )
            _write_entries(
                ws,
                base_row + 2,
                symbol_col,
                day_payload.afternoon,
                block_top=base_row,
                proto_styles=proto_styles_body,
            )
            _write_entries(
                ws,
                base_row + 4,
                symbol_col,
                day_payload.night,
                block_top=base_row,
                proto_styles=proto_styles_body,
            )

    for br in sorted(set(base_rows.values())):
        _normalize_block_chinese_numerals(
            ws, br, 1, DETAIL_TEMPLATE_SUMMARY_BEGIN_COL - 1
        )

    _refresh_detail_side_summary_formulas(ws)

    _ensure_onsheet_block_total_header(ws)
    lo_letter = get_column_letter(DETAIL_SUM_COL_START)
    hi_letter = get_column_letter(DETAIL_TEMPLATE_SUMMARY_BEGIN_COL - 1)
    for br in sorted(set(base_rows.values())):
        c_tot = ws.cell(br, DETAIL_ONSHEET_BLOCK_TOTAL_COL)
        bottom = br + DETAIL_PERSON_BLOCK_ROWS - 1
        c_tot.value = f"=SUM({lo_letter}{br}:{hi_letter}{bottom})"
        c_tot.number_format = DETAIL_BLOCK_TOTAL_NUMBER_DISPLAY

    return TemplateWriteResult(
        matched_employee_count=matched,
        unmatched_employee_names=sorted(unmatched),
    )


def _reset_sheet_rows(ws, headers: list[str]) -> None:
    ws.delete_rows(1, ws.max_row)
    ws.append(headers)


_MONTHLY_LINKABLE_METRICS: frozenset[str] = frozenset(
    {"正常上班", "平常加班", "星期天加班", "请假", "旷工", "迟到", "早退", "警告"}
)


def _excel_quoted_sheet(title: str) -> str:
    return "'" + str(title).replace("'", "''") + "'"


def _monthly_sheet_is_roster_layout(ws) -> bool:
    """识别「序号 + 图示/姓名 + 指标列」类月度模板（勿整表清空）。"""
    v = ws.cell(1, 1).value
    t = unicodedata.normalize("NFKC", _plain_cell_text(v)).replace(" ", "").replace("\n", "")
    return "序号" in t


def _scan_monthly_roster_header_row(ws, header_row: int = 1) -> tuple[int, dict[str, int]]:
    """返回 (姓名列号, 指标列号→逻辑键)。姓名列默认 2（A 为序号）。"""
    name_col = 2
    metric_cols: dict[str, int] = {}
    max_c = min(ws.max_column or 20, 40)
    for c in range(1, max_c + 1):
        raw = ws.cell(header_row, c).value
        t = unicodedata.normalize("NFKC", _plain_cell_text(raw)).replace(" ", "").replace("\n", "")
        if not t:
            continue
        if c == 1 and "序号" in t:
            continue
        if ("姓名" in t or "员工" in t) and "正常" not in t and "加班" not in t and "请假" not in t:
            name_col = c
            continue
        for key, needles in _DETAIL_MONTH_LINK_RULES:
            if key not in _MONTHLY_LINKABLE_METRICS:
                continue
            if any(n in t for n in needles):
                metric_cols[key] = c
                break
    return name_col, metric_cols


def _formula_monthly_detail_metric(
    detail_title: str,
    sumif_col_letter: str,
    name_col: int,
    ridx: int,
    base_row: int | None,
) -> str:
    """引用明细表侧栏 SUMIF 列；若未知块首行则用 MATCH(姓名, 明细!C:C) 定位。"""
    q = _excel_quoted_sheet(detail_title)
    nlet = get_column_letter(name_col)
    nm = f"{nlet}{ridx}"
    col = sumif_col_letter
    if base_row is not None:
        return f"={q}!{col}{base_row}"
    return (
        f"=IFERROR(INDEX({q}!{col}:{col},MATCH(TRIM({nm}),{q}!$C:$C,0)),\"\")"
    )


def write_analysis_sheet(workbook, rows: list[dict[str, object]]) -> None:
    ws = workbook["钉钉解析"] if "钉钉解析" in workbook.sheetnames else workbook.create_sheet("钉钉解析")
    headers = [
        "姓名",
        "考勤组",
        "部门",
        "日期",
        "班次",
        "打卡时间",
        "正班工时",
        "平常加班",
        "星期天加班",
        "请假工时",
        "旷工工时",
        "迟到次数",
        "早退次数",
        "备注",
    ]
    _reset_sheet_rows(ws, headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])


def write_monthly_sheet(
    workbook,
    rows: list[dict[str, object]],
    *,
    link_detail_side_totals: bool = True,
) -> None:
    """写入「月度统计」工作表。

    ``link_detail_side_totals`` 为真时，凡能在明细 BR..CC 表头识别到的指标列，均写公式引用
    明细表每人块首行侧栏 ``SUMIF`` 结果（与 ``_refresh_detail_side_summary_formulas`` 一致），
    手改明细符号或数值后汇总会随 Excel 重算更新。

    - 若 A1 为「序号」类排版模板，则**不清空表头样式**，仅在识别到的指标列写公式。
    - 若已知姓名在明细块首行的行号，写 ``='明细'!BRn``；否则用 ``INDEX/MATCH(姓名, 明细!C:C)`` 定位块首行。
    - 未识别到侧栏列的指标（如模板无「警告」列）仍写入聚合静态值。
    """
    ws = workbook["月度统计"] if "月度统计" in workbook.sheetnames else workbook.create_sheet("月度统计")
    headers = [
        "姓名",
        "考勤组",
        "部门",
        "工号",
        "正常上班",
        "平常加班",
        "星期天加班",
        "请假",
        "旷工",
        "迟到",
        "早退",
        "警告",
    ]

    detail_ws = workbook["明细"] if "明细" in workbook.sheetnames else None
    link_map: dict[str, int] = {}
    name_to_base: dict[str, int] = {}
    if link_detail_side_totals and detail_ws is not None:
        link_map = _detail_side_month_link_column_map(detail_ws)
        name_to_base = find_template_base_rows(detail_ws)

    detail_title = str(detail_ws.title) if detail_ws is not None else "明细"

    use_roster = _monthly_sheet_is_roster_layout(ws)
    name_col_roster, metric_cols_roster = (2, {})
    if use_roster:
        name_col_roster, metric_cols_roster = _scan_monthly_roster_header_row(ws, 1)
        if not metric_cols_roster:
            use_roster = False

    if use_roster:
        data_end = len(rows) + 1
        for r in range(data_end + 1, (ws.max_row or data_end) + 1):
            for c in metric_cols_roster.values():
                ws.cell(r, c).value = None
        h1 = ws.cell(1, 1).value
        h1t = unicodedata.normalize("NFKC", _plain_cell_text(h1)).replace(" ", "")
        use_seq_formula = "序号" in h1t
        for ridx, row in enumerate(rows, start=2):
            if use_seq_formula:
                c_seq = ws.cell(ridx, 1)
                c_seq.value = f"=ROW()-1"
                _force_arabic_number_format(c_seq, None)
            name_key = str(row.get("姓名", "") or "").strip()
            ws.cell(ridx, name_col_roster).value = name_key or None
            base_r = name_to_base.get(name_key) if name_key else None
            for key, cidx in metric_cols_roster.items():
                cell = ws.cell(ridx, cidx)
                dcol = link_map.get(key) if key in _MONTHLY_LINKABLE_METRICS else None
                if link_detail_side_totals and detail_ws is not None and dcol is not None:
                    letter = get_column_letter(int(dcol))
                    cell.value = _formula_monthly_detail_metric(
                        detail_title, letter, name_col_roster, ridx, base_r
                    )
                    _force_arabic_number_format(cell, None)
                else:
                    cell.value = row.get(key, "")
                    v = row.get(key)
                    if isinstance(v, (int, float)) and not isinstance(v, bool):
                        _force_arabic_number_format(cell, float(v))
        return

    _reset_sheet_rows(ws, headers)

    for ridx, row in enumerate(rows, start=2):
        name_key = str(row.get("姓名", "") or "").strip()
        base_r = name_to_base.get(name_key) if name_key else None
        for cidx, h in enumerate(headers, start=1):
            cell = ws.cell(ridx, cidx)
            dcol = link_map.get(h) if h in _MONTHLY_LINKABLE_METRICS else None
            if link_detail_side_totals and detail_ws is not None and dcol is not None:
                letter = get_column_letter(int(dcol))
                cell.value = _formula_monthly_detail_metric(
                    detail_title, letter, 1, ridx, base_r
                )
                _force_arabic_number_format(cell, None)
            else:
                cell.value = row.get(h, "")
                if h in {"请假", "旷工", "迟到", "早退"}:
                    v = row.get(h)
                    if isinstance(v, (int, float)) and not isinstance(v, bool):
                        _force_arabic_number_format(cell, float(v))
